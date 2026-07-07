# -*- coding: utf-8 -*-
"""
schema.json 으로 입력용 엑셀 양식을 생성한다.

생성되는 워크북: KOCARC_입력양식.xlsx
  - [사용법]   : 간단 안내
  - [환자목록] : 환자 1명 = 1행. 환자키 + 환자등록(시작) 입력 칸
  - [영역별 시트] : 공통영역, 예방역학 ... 각 시트는 환자키로 연결
  - [코드북]   : 코드값을 갖는 모든 항목의 (코드 = 의미) 사전

각 영역 시트 구조:
  1행 = 한글 라벨(질문)
  2행 = 필드명(서버 키)  ← 봇이 읽는 줄. 수정 금지.
  3행~ = 환자별 데이터 (A열 = 환자키)
"""
import os
import re
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

try:
    from example_row import EXAMPLES  # 3행 예시값(KOCARC_.xlsx 3행 기반)
except Exception:
    EXAMPLES = {}

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "KOCARC_입력양식.xlsx")

# 항목별 도움말(셀 메모). 내용은 field_notes.py 에서 사용자가 직접 채운다.
try:
    from field_notes import NOTES as FIELD_NOTES
except ImportError:
    FIELD_NOTES = {}

# 같은 질문의 단일-선택 체크박스(부/모/형제 ...)들을 드롭다운 1칸으로 묶을 때,
# 2행(필드명 줄)에 적는 마커. 봇이 이 마커를 보고 라벨->필드명 매칭을 처리한다.
# kocarc_bot.GROUP_PREFIX 와 반드시 동일해야 함.
GROUP_PREFIX = "@GROUP:"

# 조건부 라디오→라디오 트리(부모 선택에 따라 자식 라디오 활성)를 드롭다운 1칸으로
# 합칠 때 2행에 적는 마커. kocarc_bot.TREE_PREFIX 와 반드시 동일해야 함.
# 형식: @TREE:부모필드|부모값=자식필드;부모값=자식필드
TREE_PREFIX = "@TREE:"
# 잎(드롭다운 항목) 라벨 구분자: "부모라벨 - 자식라벨"
TREE_SEP = " - "

# 날짜+시+분 3칸을 1칸으로 합칠 때 2행에 적는 마커 (봇이 파싱해 3필드로 분해).
# 형식: @DT:날짜필드,시필드,분필드
DT_PREFIX = "@DT:"

# 생년월일 년/월/일 셀렉트 3칸을 1칸으로 합칠 때 2행에 적는 마커. 셀엔 'YYYY-MM-DD'.
# 형식: @DOB:년필드,월필드,일필드
DOB_PREFIX = "@DOB:"

# 값칸 + '미상'(_UK) 체크박스를 1칸으로 합칠 때 2행에 적는 마커.
# 값칸에 '미상'이라 적으면 봇이 체크박스를 누른다. 형식: @VUK:값필드,미상필드
VUK_PREFIX = "@VUK:"


def resource_path(name):
    """읽기전용 동봉 파일(schema.json). PyInstaller 번들 대응."""
    base = getattr(sys, "_MEIPASS", BASE)
    return os.path.join(base, name)

# 엑셀 시트 이름은 31자 제한 + 일부 문자 불가
SHEET_NAMES = {
    "common": "공통영역",
    "prevent": "예방역학",
    "community": "지역사회",
    "relief": "구급단계",
    "in_hosp": "병원단계",
    "alive_after": "소생후단계",
    "heart": "심장검사",
    "y_child": "소아소생술",
    "comment": "CommentLog",
}

HEADER_FILL = PatternFill("solid", fgColor="6BA68C")
NAME_FILL = PatternFill("solid", fgColor="EAF1ED")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
SECTION_FILL = PatternFill("solid", fgColor="D9E8DF")

# 입력칸(데이터 행) 공통 서식: 폰트 10, 가운데정렬, 셀에 맞춤(축소).
INPUT_FONT = Font(size=10)
INPUT_ALIGN = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
NAME_FONT = Font(color="888888", size=8)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def load_schema(schema_path=None):
    path = schema_path or resource_path("schema.json")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# 필드명 꼬리 -> 단위 (쪼개진 칸 구분용)
NAME_UNIT = [("_YY", "년"), ("_MM", "월"), ("_DD", "일"),
             ("_HOUR", "시"), ("_MIN", "분"), ("_DATE", "날짜")]
SHORT_UNITS = ("년", "월", "일", "시", "분", "세", "초")
# 라벨에 붙일 측정 단위 (용량/길이/무게 등) — 예: '에피네프린 총 투여 용량 (mg)'
MEASURE_UNITS = ("mg", "IU", "mL", "ml", "L", "g", "kg", "cc", "mmHg", "cm", "mm", "%")


# 이 단어가 든 구획명은 컬럼 라벨 접두어로 붙이지 않는다(검사 패널처럼 44칸이 다 길어짐).
NO_PREFIX_CONTEXTS_KW = ("혈액검사",)


def with_context(label, ctx):
    """질병명/구획 같은 상위 컨텍스트를 라벨 앞에 붙인다. (예: '고혈압 - 치료')
    단, 검사 패널 등 너무 긴 구획명은 접두어로 붙이지 않는다(라벨만 사용)."""
    ctx = (ctx or "").strip()
    if not ctx or ctx in label or any(kw in ctx for kw in NO_PREFIX_CONTEXTS_KW):
        return label
    return f"{ctx} - {label}"


# 자동 라벨이 헷갈리는 특정 필드의 표시 라벨 직접 지정.
LABEL_OVERRIDES = {
    "HOSP_RESULT_OUT": "병원치료결과 (퇴원처)",  # 자택/타병원전원/호스피스/요양시설
    "EPINE_TOT": "에피네프린 용량 (응급실 CPR 시행 환자)",
    "EPINE_ROSC": "에피네프린 용량 (Sustained ROSC 환자)",
    "MULTI_MOVE": "출동 종류",  # 원본 라벨 '다중출동' → 단일/다중 종류 고르는 칸이라 헷갈림
}

# 배타적(택1) 체크박스 그룹의 group_prefix — 복수선택 안내 대신 '택1' 안내를 붙인다.
SINGLE_SELECT_GROUPS = {"EPINE"}  # 에피네프린 미상/미사용 (사이트에서 하나 고르면 나머지 비활성)
# 그룹 컬럼 라벨 직접 지정 (group_prefix 기준)
GROUP_LABEL_OVERRIDES = {"EPINE": "에피네프린 용량 (미상/미사용)"}


def display_label(f):
    """사람이 읽기 쉬운 컬럼 라벨(상위 컨텍스트 포함)."""
    if f["name"] in LABEL_OVERRIDES:
        return LABEL_OVERRIDES[f["name"]]
    return with_context(_display_core(f), f.get("context"))


def _display_core(f):
    """컬럼 라벨 본문(컨텍스트 제외).
    - 쪼개진 일시/생년월일: (년)(월)(일)(날짜)(시)(분)
    - 단일 체크박스/라디오: 그 칸의 의미(부/모/미상/ND 등)를 붙임
    - 단위/미상 보조칸: (단위)(미상)"""
    base = f.get("label") or f["name"]
    up = f["name"].upper()
    # 1) 날짜/시각/생년월일 분할 단위
    for sfx, u in NAME_UNIT:
        if up.endswith(sfx):
            return f"{base} ({u})"
    # 2) 단일 체크박스/라디오: 옆에 적힌 의미를 붙임
    if f["type"] in ("radio", "checkbox") and len(f.get("choices") or []) == 1:
        lab = (f["choices"][0].get("label") or "").strip()
        if lab:
            return f"{base} ({lab})"
    # 3) 단위/미상/기타 텍스트칸 (뒤에 번호가 붙어도: _ETC1 등)
    if re.search(r"_UNIT\d*$", up):
        return f"{base} (단위)"
    if re.search(r"_UK\d*$", up):
        return f"{base} (미상)"
    if re.search(r"_ETC\d*$", up):
        return f"{base} (기타)"
    # 4) 짧은 꼬리표(시/분/년/월/일/세) + 측정 단위(mg/IU 등)
    s = (f.get("suffix") or "").strip()
    if s in SHORT_UNITS or s in MEASURE_UNITS:
        return f"{base} ({s})"
    return base


def labeled_fields(fields):
    """user_entry 필드들에 표시 라벨을 부여하고, 같은 라벨은 [필드명]으로 구분."""
    out, seen = [], {}
    for f in fields:
        if not f["user_entry"]:
            continue
        d = display_label(f)
        if d in seen:
            seen[d] += 1
            d = f"{d} [{f['name']}]"
        else:
            seen[d] = 1
        out.append((f, d))
    return out


class ListSheet:
    """드롭다운 목록 저장용 시트.
    인라인 목록("a,b,c")은 라벨에 콤마/따옴표가 있거나 255자를 넘으면 못 쓴다.
    그런 경우 라벨을 이 시트의 한 열에 세로로 적고, 그 범위를 참조하는 드롭다운을 만든다.
    사용자가 나중에 항목을 고칠 수 있게 보이는 시트로 두고, 1행에 어떤 항목인지 제목을 단다."""
    NAME = "_목록"

    def __init__(self, wb):
        self.ws = wb.create_sheet(self.NAME)
        self.ws.cell(row=1, column=1)  # 시트 존재 보장
        self.ws.freeze_panes = "A2"
        self.col = 0

    def add_range(self, labels, title=None):
        self.col += 1
        letter = get_column_letter(self.col)
        h = self.ws.cell(row=1, column=self.col, value=title or "목록")
        h.font = Font(bold=True, size=9, color="666666")
        for i, lab in enumerate(labels, start=1):
            self.ws.cell(row=i + 1, column=self.col, value=lab)
        self.ws.column_dimensions[letter].width = 24
        # 1행은 제목, 실제 항목은 2행부터
        return f"'{self.NAME}'!${letter}$2:${letter}${1 + len(labels)}"


def make_dv(list_ctx, labels, show_error=False, error_text=None,
            error_title=None, title=None):
    """라벨 목록 → 드롭다운. 인라인이 안전하면 인라인, 아니면(콤마/따옴표/255초과)
    _목록 시트 범위 참조로 대체(list_ctx 없으면 None)."""
    labels = [x for x in labels if x]
    if not labels:
        return None
    inline = '"' + ",".join(labels) + '"'
    if not any(("," in x or '"' in x) for x in labels) and len(inline) <= 255:
        formula = inline
    elif list_ctx is not None:
        formula = list_ctx.add_range(labels, title)
    else:
        return None
    dv = DataValidation(type="list", formula1=formula,
                        allow_blank=True, showErrorMessage=show_error)
    if show_error and error_text:
        dv.error = error_text
        dv.errorTitle = error_title
    return dv


def dv_for_field(f, list_ctx=None, title=None):
    """선택지가 적은 코드필드에 한해 드롭다운(라벨 기준) 생성."""
    labels = []
    if f["type"] in ("radio", "checkbox"):
        labels = [c["label"] for c in f["choices"] if c["label"]]
    elif f["type"] == "select":
        # 자리표시(placeholder) 옵션 제외
        labels = [o["label"] for o in f["options"]
                  if o["value"] not in ("", "XXXX", "XX", "--", "XX:-", "--:-")
                  and o["label"] not in ("Y", "M", "D", "-")]
    if not labels or len(labels) > 20:
        return None
    return make_dv(list_ctx, labels, show_error=True,
                   error_text="코드북의 값 중에서 선택하세요", error_title="값 확인",
                   title=title or f["name"])


def is_single_cb(f):
    """value 없는 '해당시 체크' 단일 체크박스(부/모/형제 같은 칸) 여부."""
    return (f["type"] == "checkbox"
            and len(f.get("choices") or []) == 1
            and bool((f["choices"][0].get("label") or "").strip()))


def group_prefix(name):
    """그룹 공통 접두어. 끝 숫자(HTNTX1->HTNTX) 또는 끝 _세그먼트(FHX_CA_F->FHX_CA)를 떼어낸다."""
    p = re.sub(r"\d+$", "", name)
    if p == name and "_" in name:
        p = name.rsplit("_", 1)[0]
    return p


def grouped_columns(fields):
    """user_entry 필드를 '컬럼 단위'로 묶는다.
    같은 질문(label)·같은 접두어의 단일 체크박스가 2개 이상 연속하면 하나의 그룹으로.
    반환 항목: ("single", field) 또는 ("group", 라벨, [member_fields])"""
    uf = [f for f in fields if f["user_entry"]]
    cols, i = [], 0
    while i < len(uf):
        f = uf[i]
        if is_single_cb(f):
            lab, pre = f["label"], group_prefix(f["name"])
            members, deferred, j = [], [], i
            while j < len(uf):
                g = uf[j]
                same = (g.get("label") == lab and group_prefix(g["name"]) == pre)
                if is_single_cb(g) and same:
                    members.append(g)        # 같은 질문의 단일 체크박스 → 그룹 멤버
                elif same:
                    deferred.append(g)        # 같은 질문의 자유입력(_ETC 등) → 그룹 뒤로
                else:
                    break                     # 다른 질문이면 그룹 종료
                j += 1
            if len(members) >= 2:
                cols.append(("group", lab or f["name"], members))
                for d in deferred:            # 중간에 낀 _ETC 등은 그룹 바로 뒤 컬럼으로
                    cols.append(("single", d))
                i = j
                continue
        cols.append(("single", f))
        i += 1
    return cols


def dv_for_group(members, list_ctx=None, title=None):
    """그룹 드롭다운: 멤버 라벨(부/모/...) 목록. 드롭다운을 주되,
    복수선택(쉼표 입력)을 막지 않도록 검증 오류는 띄우지 않는다."""
    labels = [(m["choices"][0].get("label") or "").strip() for m in members]
    labels = [x for x in labels if x]
    if not labels or len(labels) > 30:
        return None
    return make_dv(list_ctx, labels, show_error=False, title=title)


def _leaf_paths(byname, field):
    """필드의 드롭다운 잎 경로들. 자식이 또 부모(tree)면 손자까지 재귀.
    각 잎 = 라벨 리스트(예: ['Medical (내과적원인)', 'Presumed Cardiac', 'Arrythmia/other'])."""
    tree = field.get("tree") or {}
    out = []
    for c in field.get("choices", []):
        pv, plab = c.get("value", ""), (c.get("label") or "").strip()
        if not plab:
            continue
        child = byname.get(tree.get(pv))
        if child:
            for sub in _leaf_paths(byname, child):
                out.append([plab] + sub)
        else:
            out.append([plab])
    return out


def tree_leaves(byname, parent):
    """트리 부모 필드 → (드롭다운 잎 라벨 목록, @TREE 마커 문자열).
    잎은 '부모 - 자식(- 손자)' 로 펼침. 마커는 직속 자식만 인코딩(부모값=자식필드);
    더 깊은 단계는 봇이 schema.json 의 자식필드 tree 로 복원한다.
    ponytail: 잎 라벨에 ' - '가 든 선택지가 생기면 봇 분리가 깨짐(현 스키마엔 없음)."""
    tree = parent.get("tree") or {}
    leaves = [TREE_SEP.join(path) for path in _leaf_paths(byname, parent)]
    marker = TREE_PREFIX + parent["name"] + "|" + ";".join(
        f"{pv}={cn}" for pv, cn in tree.items())
    return leaves, marker


def dv_from_labels(labels, list_ctx=None, title=None):
    """라벨 목록 → 드롭다운(검증 오류 없음: 통합값이라 자유 확인)."""
    labels = [x for x in labels if x]
    if not labels or len(labels) > 40:
        return None
    return make_dv(list_ctx, labels, show_error=False, title=title)


def _is_date(f):
    return f["widget"] == "date" or f["name"].upper().endswith("_DATE")


def _is_hour(f):
    return f["widget"] == "time_hour" or f["name"].upper().endswith("_HOUR")


def _is_min(f):
    return f["widget"] == "time_min" or f["name"].upper().endswith("_MIN")


def merge_datetime(specs):
    """연속된 (날짜, 시, 분) 단일칸 3개를 하나의 datetime 칸으로 합친다.
    이름 접두어가 같아야만 합쳐 오짝을 막는다. (날짜 이름은 _DATE 로 끝남)
    바로 뒤에 같은 접두어의 '미상'(_UK) 단일 체크박스가 있으면 같은 칸에 흡수해
    ("datetime", 날짜, 시, 분, 미상) 5-튜플로 만든다(엑셀칸에 '미상' 입력 흡수)."""
    out, i = [], 0
    while i < len(specs):
        s = specs[i]
        # 날짜 이름은 _DATE 또는 _DATE+번호(SSEP_DATE1)로 끝남 → 접두어(pre) 추출
        m = (re.match(r"^(.*)_DATE\d*$", s[1]["name"], re.I)
             if (s[0] == "single" and _is_date(s[1])) else None)
        if (m and i + 2 < len(specs)
                and specs[i + 1][0] == "single" and _is_hour(specs[i + 1][1])
                and specs[i + 2][0] == "single" and _is_min(specs[i + 2][1])):
            pre = m.group(1)
            if (specs[i + 1][1]["name"].startswith(pre)
                    and specs[i + 2][1]["name"].startswith(pre)):
                merged = ["datetime", s[1], specs[i + 1][1], specs[i + 2][1]]
                step = 3
                nxt = specs[i + 3] if i + 3 < len(specs) else None
                # 미상 체크박스: _UK 또는 _UK+번호(SSEP_DATE_UK1)
                if (nxt and nxt[0] == "single" and is_single_cb(nxt[1])
                        and re.search(r"_UK\d*$", nxt[1]["name"], re.I)
                        and nxt[1]["name"].startswith(pre)):
                    merged.append(nxt[1])  # 미상 체크박스 흡수
                    step = 4
                out.append(tuple(merged))
                i += step
                continue
        out.append(s)
        i += 1
    return out


# 값칸 뒤 '미측정/미상' 플래그 체크박스로 인정할 안전 키워드(그 라벨이 이 안에 있어야 합침).
VALUK_KEYWORDS = {"ND", "미상", "미측정", "모름"}

# 이름 규칙(_UK/_YN)으로 못 잡는 값칸+플래그 수동 페어링. {값필드: 플래그체크박스필드}
MANUAL_VALUK = {
    "DEFIB_CNT": "DEFIB_UK",  # 제세동 횟수 + 미상 (이름이 DEFIB_CNT_UK 가 아님)
}


def merge_value_uk(specs):
    """값 입력칸(텍스트/숫자) 바로 뒤에 '{이름}_UK'(미상) 또는 '{이름}_YN'(ND) 플래그
    체크박스가 오면 한 칸으로 합친다. 사용자가 값칸에 그 키워드(미상/ND)를 적으면 봇이
    체크박스를 누른다. 키워드는 체크박스 라벨에서 가져와 마커에 담는다.
    라디오/체크박스/셀렉트 값칸은 대상 아님(자유입력 값칸만)."""
    out, i = [], 0
    while i < len(specs):
        s = specs[i]
        nxt = specs[i + 1] if i + 1 < len(specs) else None
        merged = None
        if (s[0] == "single" and s[1]["type"] not in ("radio", "checkbox", "select")
                and nxt and nxt[0] == "single" and is_single_cb(nxt[1])):
            vn, cn = s[1]["name"], nxt[1]["name"]
            kw = (nxt[1]["choices"][0].get("label") or "").strip()
            kw = kw.split()[0] if kw else kw  # 'ND -' 같은 구분자 꼬리 제거
            paired = cn in (vn + "_UK", vn + "_YN") or MANUAL_VALUK.get(vn) == cn
            if paired and kw in VALUK_KEYWORDS:
                merged = ("valuk", s[1], nxt[1], kw)
        if merged:
            out.append(merged)
            i += 2
            continue
        out.append(s)
        i += 1
    return out


def merge_dob(specs):
    """연속된 (년, 월, 일) 셀렉트 3개({접두}_YY/_MM/_DD)를 생년월일 한 칸으로 합친다.
    사용자가 'YYYY-MM-DD'로 적으면 봇이 세 드롭다운으로 분해해 넣는다."""
    out, i = [], 0
    while i < len(specs):
        s = specs[i]
        m = (re.match(r"^(.*)_YY$", s[1]["name"], re.I)
             if s[0] == "single" else None)
        if (m and i + 2 < len(specs)
                and specs[i + 1][0] == "single" and specs[i + 2][0] == "single"):
            pre = m.group(1)
            if (specs[i + 1][1]["name"].upper() == (pre + "_MM").upper()
                    and specs[i + 2][1]["name"].upper() == (pre + "_DD").upper()):
                out.append(("dob", s[1], specs[i + 1][1], specs[i + 2][1]))
                i += 3
                continue
        out.append(s)
        i += 1
    return out


def write_columns(ws, fields, n_rows, list_ctx=None):
    """area/patient 시트 공통: B열부터 컬럼(단일/그룹/일시)을 쓰고 환자키를 채운다.
    반환: (입력 칸 개수, {필드명: 컬럼번호})  ← 컬럼맵은 조건부 서식에서 사용."""
    seen, col = {}, 2
    colmap = {}
    byname = {f["name"]: f for f in fields}
    for spec in merge_dob(merge_value_uk(merge_datetime(grouped_columns(fields)))):
        example = ""  # 3행(회색 예시행)에 넣을 예시값 — 봇은 A열 빈 이 행을 무시
        if spec[0] == "valuk":
            _, fv, fuk, kw = spec
            colmap[fv["name"]] = col
            colmap[fuk["name"]] = col
            disp = display_label(fv)  # 값칸 라벨(단위 포함)
            if disp in seen:
                seen[disp] += 1
                disp = f"{disp} [{fv['name']}]"
            else:
                seen[disp] = 1
            marker = VUK_PREFIX + ",".join([fv["name"], fuk["name"], kw])
            _write_header(ws, col, disp, marker)
            dv = DataValidation(showInputMessage=True, promptTitle="입력",
                                prompt=f"값 입력. 모르거나 미측정이면 '{kw}' 입력")
        elif spec[0] == "datetime":
            dts = spec[1:]  # (날짜, 시, 분) 또는 (날짜, 시, 분, 미상체크박스)
            for x in dts:
                colmap[x["name"]] = col  # 합친 칸을 가리킴(조건부서식용)
            fd = dts[0]
            disp = with_context(fd.get("label") or fd["name"], fd.get("context"))
            if disp in seen:
                # 값칸과 라벨이 겹치는 일시칸 → 필드명 대신 (시행일시)로 구분
                cand = f"{disp} (시행일시)"
                disp = cand if cand not in seen else f"{disp} [{fd['name']}]"
            seen[disp] = seen.get(disp, 0) + 1
            marker = DT_PREFIX + ",".join(x["name"] for x in dts)
            _write_header(ws, col, disp, marker)
            # 12자리 숫자로 입력하면 셀에 2026-06-15 05:02 로 자동 표시(저장값은 숫자).
            for r in range(4, 4 + n_rows):
                ws.cell(row=r, column=col).number_format = "0000-00-00 00\\:00"
            prompt = "12자리 숫자로 입력하면 2026-06-15 05:02 로 자동 표시 (예: 202606150502)"
            if len(dts) == 4:  # 미상 흡수됨
                prompt += "  ·  모르면 '미상'"
            dv = DataValidation(showInputMessage=True, promptTitle="일시 입력형식",
                                prompt=prompt)
            example = "2026-06-15 09:30"
        elif spec[0] == "dob":
            parts = spec[1:]  # (년, 월, 일 셀렉트)
            for x in parts:
                colmap[x["name"]] = col
            fd = parts[0]
            disp = fd.get("label") or fd["name"]  # '생년월일'
            if disp in seen:
                disp = f"{disp} [{fd['name']}]"
            seen[disp] = seen.get(disp, 0) + 1
            marker = DOB_PREFIX + ",".join(x["name"] for x in parts)
            _write_header(ws, col, disp, marker)
            # 8자리 숫자로 입력하면 셀에 1970-05-15 로 자동 표시(저장값은 숫자 그대로).
            for r in range(4, 4 + n_rows):
                ws.cell(row=r, column=col).number_format = "0000-00-00"
            dv = DataValidation(showInputMessage=True, promptTitle="생년월일 입력형식",
                                prompt="8자리 숫자로 입력하면 1970-05-15 로 자동 표시됩니다.")
            example = "1970-05-15"
        elif spec[0] == "single":
            f = spec[1]
            colmap[f["name"]] = col
            disp = display_label(f)
            if disp in seen:
                seen[disp] += 1
                disp = f"{disp} [{f['name']}]"
            else:
                seen[disp] = 1
            if f.get("tree"):
                leaves, marker = tree_leaves(byname, f)
                _write_header(ws, col, disp, marker)
                dv = dv_from_labels(leaves, list_ctx, title=disp)
            else:
                _write_header(ws, col, disp, f["name"])
                dv = dv_for_field(f, list_ctx, title=disp)
        else:
            _, glabel, members = spec
            colmap[members[0]["name"]] = col  # 그룹 컬럼도 조건부서식용으로 기록
            gp = group_prefix(members[0]["name"])
            single = gp in SINGLE_SELECT_GROUPS
            if gp in GROUP_LABEL_OVERRIDES:
                disp = GROUP_LABEL_OVERRIDES[gp]
            else:
                disp = with_context(glabel, members[0].get("context"))
            if disp in seen:
                seen[disp] += 1
                # 같은 질문의 라디오(예/아니오/미상) 옆 세부 체크박스 묶음 → "(상세)"
                alt = f"{disp} (상세)"
                if alt in seen:
                    alt = f"{disp} [{gp}]"
                disp = alt
                seen[disp] = seen.get(disp, 0) + 1
            else:
                seen[disp] = 1
            name_text = GROUP_PREFIX + ",".join(m["name"] for m in members)
            _write_header(ws, col, disp, name_text)
            dv = dv_for_group(members, list_ctx, title=disp)
            if dv is not None:
                dv.showInputMessage = True
                if single:  # 배타적 그룹 → 택1 안내
                    opts = " / ".join((m["choices"][0].get("label") or "").strip()
                                      for m in members)
                    dv.promptTitle = "택1 (하나만 선택)"
                    dv.prompt = f"하나만 선택하세요: {opts}"
                else:       # 복수선택 그룹 → 쉼표 안내
                    ex = ", ".join(
                        (m["choices"][0].get("label") or "").strip()
                        for m in members[:2] if (m["choices"][0].get("label") or "").strip())
                    dv.promptTitle = "복수 선택 가능"
                    dv.prompt = ("해당하는 것을 모두 쉼표(,)로 구분해 입력하세요."
                                 + (f"  예: {ex}" if ex else ""))
        if dv is not None:
            ws.add_data_validation(dv)
            colL = get_column_letter(col)
            dv.add(f"{colL}3:{colL}{3 + n_rows}")
        # 3행 = 회색 예시행(입력 무시).
        # 이 시트가 예시맵(KOCARC_.xlsx 3행)에 있으면 그 값을 그대로 사용(빈칸 포함).
        ex_map = EXAMPLES.get(ws.title)
        marker = ws.cell(row=2, column=col).value  # _write_header 가 쓴 2행 마커
        if ex_map is not None and marker in ex_map:
            example = ex_map[marker]
        # 예시맵에 없는 시트만 드롭다운 첫 항목을 예시로 자동 사용.
        elif not example and dv is not None and dv.type == "list" \
                and dv.formula1 and dv.formula1.startswith('"'):
            example = dv.formula1.strip('"').split(",")[0]
        if example not in (None, ""):
            exc = ws.cell(row=3, column=col, value=example)
            exc.font = Font(italic=True, color="A6A6A6", size=9)
        col += 1
    for i in range(n_rows):
        ws.cell(row=4 + i, column=1, value=i + 1)   # 환자키: 데이터는 4행부터
    # 예시행(3행)+데이터 칸: 테두리 + 가운데정렬 + 셀에맞춤. 입력행 폰트 10.
    for r in range(3, 4 + n_rows):
        for cc in range(1, col):
            cell = ws.cell(row=r, column=cc)
            cell.border = BORDER
            cell.alignment = INPUT_ALIGN
            if r >= 4:                       # 예시행(3)의 9pt 이탤릭 폰트는 유지
                cell.font = INPUT_FONT
    ws.freeze_panes = "B4"
    return col - 2, colmap


# 조건부 회색 규칙 (안내용 표시 — 입력 잠금이 아님).
# 부모 통합칸 값이 아래 '흰색(입력) 조건'을 만족하지 않으면 대상 칸을 회색으로.
# 통합칸 값은 "부모라벨 - 자식라벨" 문자열이므로 그 문자열을 그대로 비교한다.
# 조건부 서식(dxf) 채우기는 fgColor 만으로는 Excel 이 렌더링하지 않는다.
# start_color+end_color+fill_type 로 지정해야 실제로 색이 보인다.
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def add_xref_helper(ws, src_sheet, src_col, n_rows, at_col):
    """다른 시트(src_sheet)의 src_col 값을 이 시트 at_col 에 미러링하는 '숨김 헬퍼 열'.
    조건부서식은 교차시트 참조(`시트!$B3`·INDIRECT 포함)를 환경에 따라 렌더링하지 않으므로,
    같은 시트 헬퍼 열을 참조하게 만든다. 일반 셀 수식의 교차시트 참조는 항상 작동한다.
    2행(필드명)을 비워 봇이 이 열을 무시하게 한다(read_sheet_rows: 2행 없으면 skip).
    반환: 헬퍼 열 문자."""
    L = get_column_letter(at_col)
    ws.cell(row=1, column=at_col, value="↩자동참조(수정·삭제 금지)")
    for r in range(3, 4 + n_rows):                       # 같은 환자=같은 행
        ref = f"'{src_sheet}'!${src_col}{r}"
        ws.cell(row=r, column=at_col, value=f'=IF({ref}="","",{ref})')  # 빈칸은 빈칸 유지
    ws.column_dimensions[L].hidden = True
    return L


def dt_serial(ref):
    """12자리 숫자(YYYYMMDDHHMM) 셀/식 → 엑셀 날짜시간 시리얼(정수1=하루). 24시간=1."""
    t = f'TEXT({ref},"000000000000")'
    return (f'DATE(LEFT({t},4),MID({t},5,2),MID({t},7,2))'
            f'+TIME(MID({t},9,2),MID({t},11,2),0)')


def add_prompt(ws, col_letter, n_rows, text, title="확인 규칙"):
    """지정 열의 데이터 칸(3행~) 입력 메시지(=셀 클릭 시 뜨는 툴팁)에 규칙/경고 사유를 붙인다.
    이미 DV(프롬프트)가 있으면 이어붙이고, 없으면 새로 만든다(칸당 DV 1개 유지)."""
    rng = f"{col_letter}3:{col_letter}{3 + n_rows}"
    for dv in ws.data_validations.dataValidation:
        if str(dv.sqref) == rng:
            dv.showInputMessage = True
            base = (dv.prompt or "").rstrip()
            if text not in base:
                dv.prompt = (base + "\n" if base else "") + text
            if not dv.promptTitle:
                dv.promptTitle = title
            return
    dv = DataValidation(showInputMessage=True, promptTitle=title, prompt=text)
    ws.add_data_validation(dv)
    dv.add(rng)


# ── @DT(12자리 일시) 규칙 헬퍼 ──────────────────────────────────────────
# 순서 위반/상한 초과 시 빨강. raw 12자리(YYYYMMDDHHMM)는 시간순 단조라 순서비교는
# 그대로, 'N분 이내' 상한은 dt_serial 로 실제 시간(1=하루)으로 바꿔 비교한다.
def _dt_order(ws, colmap, n_rows, target, ref_cell, op, prompt):
    """target(@DT) 이 ref_cell 보다 op 관계면 빨강. op '<'=target이 ref보다 이전.
    ref_cell: 같은 시트 셀('$H3') 또는 환자목록 미러 헬퍼셀. 둘 다 raw 12자리."""
    if target not in colmap:
        return
    c = get_column_letter(colmap[target])
    t = f"${c}3"
    cond = f"AND(ISNUMBER({t}),ISNUMBER({ref_cell}),{t}{op}{ref_cell})"
    ws.conditional_formatting.add(
        f"{c}3:{c}{3 + n_rows}", FormulaRule(formula=[cond], fill=WARN_FILL))
    add_prompt(ws, c, n_rows, prompt, "일시 규칙")


def _dt_window(ws, colmap, n_rows, target, ref_cell, lo_min, hi_min, prompt):
    """target 이 [ref+lo_min분, ref+hi_min분] 범위 밖이면 빨강. lo/hi None=무제한."""
    if target not in colmap:
        return
    c = get_column_letter(colmap[target])
    tS, rS = dt_serial(f"${c}3"), dt_serial(ref_cell)
    parts = []
    if lo_min is not None:
        parts.append(f"{tS}<{rS}+({lo_min})/1440")
    if hi_min is not None:
        parts.append(f"{tS}>{rS}+({hi_min})/1440")
    inner = f"OR({','.join(parts)})" if len(parts) > 1 else parts[0]
    cond = f"AND(ISNUMBER(${c}3),ISNUMBER({ref_cell}),{inner})"
    ws.conditional_formatting.add(
        f"{c}3:{c}{3 + n_rows}", FormulaRule(formula=[cond], fill=WARN_FILL))
    add_prompt(ws, c, n_rows, prompt, "일시 규칙")


def _er_helper_cell(ws, colmap, patient_colmap, n_rows):
    """환자목록 ER_DATE(내원시각)를 미러링한 숨김 헬퍼열의 3행 셀('$X3') 반환.
    CF의 교차시트 참조가 환경따라 안 먹혀 같은 시트 헬퍼열로 우회. 없으면 None.
    ws.max_column+1 에 두어 이 시트가 이미 만든 다른 헬퍼열과 겹치지 않게 한다."""
    if not (patient_colmap and "ER_DATE" in patient_colmap and colmap):
        return None
    av = get_column_letter(patient_colmap["ER_DATE"])
    h = add_xref_helper(ws, "환자목록", av, n_rows, ws.max_column + 1)
    return f"${h}3"


# 환자등록(환자목록 시트) 필수항목: 사이트 checkInput 이 빈칸이면 저장을 막는 칸.
# AGE 는 생년월일로 사이트가 자동계산 → 입력칸이 아니라 제외.
REQUIRED_PATIENT_FIELDS = ["PAT_NM", "SEX", "DOB_YY", "ER_DATE", "HOSP_CD"]


def apply_required_rules(ws, colmap, n_rows, field_names=REQUIRED_PATIENT_FIELDS, log=print):
    """필수 입력칸이 비었는데 그 행에 입력이 시작돼 있으면 빨강 경고.
    환자키는 1~N 미리 채워져 있으므로 '키 있음'으론 사용중 행을 못 가른다 →
    필수칸 중 하나라도 채워졌는지(COUNTA>0)로 판별해 빈 행은 표시하지 않는다."""
    last = 3 + n_rows
    cols, missing = [], []
    for name in field_names:
        if name in colmap:
            cols.append(get_column_letter(colmap[name]))
        else:
            missing.append(name)
    cols = list(dict.fromkeys(cols))          # @DOB/@DT 형제 필드가 같은 열이면 중복 제거
    if not cols:
        log(f"[필수항목 규칙] 대상 열을 못 찾음: {field_names}")
        return
    # 4행부터(예시행 3행은 제외 — B/D/E 예시값 때문에 이니셜칸이 오탐되는 것 방지).
    counta = "COUNTA(" + ",".join(f"${c}4" for c in cols) + ")"
    for c in cols:
        f = f'AND(${c}4="",{counta}>0)'        # 이 칸 비었고 + 행에 입력이 시작됨
        ws.conditional_formatting.add(
            f"{c}4:{c}{last}", FormulaRule(formula=[f], fill=WARN_FILL))
        add_prompt(ws, c, n_rows,
                   "⚠ 필수항목입니다. 비우면 사이트에서 저장이 막힙니다.", "필수항목")
    if missing:
        log(f"[필수항목 규칙] 입력칸 아님/제외로 건너뜀: {missing}")


def apply_grey_rules(ws, colmap, n_rows, patient_colmap=None):
    """공통영역: 심폐소생술1/2 선택에 따라 종속 일시칸을 회색 처리.
    (심폐1=ER_CPR1, 심폐2=ER_CPR2 통합칸. 값은 '부모 - 자식' 문자열.)"""
    if "ER_CPR1" not in colmap or "ER_CPR2" not in colmap:
        return
    last = 3 + n_rows
    c1 = get_column_letter(colmap["ER_CPR1"])  # 심폐소생술1 통합칸
    c2 = get_column_letter(colmap["ER_CPR2"])  # 심폐소생술2 통합칸
    c3 = get_column_letter(colmap["ER_CPR_RESULT"]) if "ER_CPR_RESULT" in colmap else None
    # (대상 필드들, 회색으로 만들 조건[= 흰색 조건의 부정], 3행 기준 수식)
    rules = [
        # 중단일시: 심폐1이 '미시행'이면 회색 (CPR 미시행 → 중단시각 없음). 미입력도 회색.
        (["ER_CPR_STOP_DATE", "ER_CPR_STOP_HOUR", "ER_CPR_STOP_MIN"],
         f'OR(${c1}3="",LEFT(${c1}3,3)="미시행")'),
        # Any ROSC 일시: 심폐2 = '시행 - Sustained ROSC없이 Any ROSC만' 일 때만 흰색.
        (["ER_ANY_ROSC_DATE", "ER_ANY_ROSC_HOUR", "ER_ANY_ROSC_MIN"],
         f'${c2}3<>"시행 - Sustained ROSC없이 Any ROSC만"'),
        # Sustained ROSC 일시: 심폐2='시행 - Sustained ROSC' 또는
        #                      심폐1='시행 20분이내 중단 - 자발순환회복' 일 때만 흰색.
        (["ER_SUS_ROSC_DATE", "ER_SUS_ROSC_HOUR", "ER_SUS_ROSC_MIN"],
         f'AND(${c2}3<>"시행 - Sustained ROSC",'
         f'${c1}3<>"시행 20분이내 중단 - 자발순환회복")'),
    ]
    c4 = get_column_letter(colmap["HOSP_RESULT"]) if "HOSP_RESULT" in colmap else None
    c5 = get_column_letter(colmap["FU6M_STAT"]) if "FU6M_STAT" in colmap else None
    if c3:
        rules += [
            # 응급실 사망일시: 결과가 '사망'(사망-사망/가망없는퇴원)일 때만 흰색.
            (["ER_DIE_DATE", "ER_DIE_HOUR", "ER_DIE_MIN"],
             f'LEFT(${c3}3,2)<>"사망"'),
            # 생존-입원/퇴원 종속칸: 결과='생존 - 입원' 또는 '생존 - 퇴원'일 때 흰색
            # (사이트 setER_CPR_RESULT_LIVE 와 동일: 입원 OR 퇴원).
            # HOSP_CPC/FU6M_DIE/F6M_CPC 는 아래에서 세부 게이팅.
            (["HOSP_AD_DATE", "HOSP_AD_HOUR", "HOSP_AD_MIN", "HOSP_AD_ROOM",
              "HOSP_AD_STAT", "HOSP_RESULT", "FU6M_STAT"],
             f'AND(${c3}3<>"생존 - 입원",${c3}3<>"생존 - 퇴원")'),
        ]
    if c4:  # 병원치료결과 하위
        rules += [
            # 병원퇴원일시: 병원치료결과='생존퇴원 - …'(트리 통합)일 때만 흰색.
            # 퇴원처(HOSP_RESULT_OUT)는 이 부모 드롭다운에 흡수됨.
            (["HOSP_OUT_DATE", "HOSP_OUT_HOUR", "HOSP_OUT_MIN"],
             f'LEFT(${c4}3,4)<>"생존퇴원"'),
            # 병원사망일시: 병원치료결과='사망퇴원'일 때만 흰색.
            (["HOSP_DIE_DATE", "HOSP_DIE_HOUR", "HOSP_DIE_MIN"],
             f'${c4}3<>"사망퇴원"'),
        ]
    if c3 and c4:
        # 병원퇴원시 신경학적상태: 결과=생존-입원/퇴원 이면서 병원치료결과≠입원중 일 때만 흰색.
        rules.append(
            (["HOSP_CPC"],
             f'OR(AND(${c3}3<>"생존 - 입원",${c3}3<>"생존 - 퇴원"),${c4}3="입원중")'))
    if c5:  # 6개월 후 생존 하위
        rules += [
            # 6개월 후 사망일시: 6개월 후 생존='사망'일 때만 흰색.
            (["FU6M_DIE_DATE", "FU6M_DIE_HOUR", "FU6M_DIE_MIN"],
             f'${c5}3<>"사망"'),
            # 6개월 후 신경학적 상태: 6개월 후 생존='생존'일 때만 흰색.
            (["F6M_CPC"], f'${c5}3<>"생존"'),
        ]
    seen_cf = set()
    for fnames, formula in rules:
        for fn in fnames:
            if fn not in colmap:
                continue
            c = get_column_letter(colmap[fn])
            if (c, formula) in seen_cf:  # 합쳐진 일시칸 중복 방지
                continue
            seen_cf.add((c, formula))
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    # 모순 경고: 심폐1이 '미시행'이 아닌데(=시행) 심폐2='미시행'이면 논리 모순 → 심폐2칸 빨강.
    warn_f = f'AND(${c2}3="미시행",${c1}3<>"",LEFT(${c1}3,3)<>"미시행")'
    ws.conditional_formatting.add(
        f"{c2}3:{c2}{last}", FormulaRule(formula=[warn_f], fill=WARN_FILL))
    # 고르기 전에도 알도록: 심폐2 칸 클릭시 안내문.
    for dv in ws.data_validations.dataValidation:
        if str(dv.sqref) == f"{c2}3:{c2}{last}":
            dv.showInputMessage = True
            dv.promptTitle = "심폐소생술2"
            dv.prompt = ("심폐소생술1이 '미시행'이 아니면(=시행) 여기서 '미시행'은 모순입니다.\n"
                         "잘못 고르면 칸이 빨갛게 표시됩니다.")
            break

    # 시각 순서 경고: 응급실/입원 일시는 등록병원 응급실 내원시각(환자목록!ER_DATE)보다
    # 이후여야 한다. 같거나 이전이면 빨강. @DT칸은 12자리 숫자(YYYYMMDDHHMM)라 숫자비교=시간비교.
    # ISNUMBER 가드: 미상(문자)·빈칸·예시행(문자열)은 자동 제외.
    # ponytail: 회색규칙과 겹치는 칸은 규칙 우선순위 미조정 — 회색은 보통 빈칸이라 실무상 충돌 드묾.
    if patient_colmap and "ER_DATE" in patient_colmap and colmap:
        av = get_column_letter(patient_colmap["ER_DATE"])
        harr = add_xref_helper(ws, "환자목록", av, n_rows, max(colmap.values()) + 1)
        arr = f"${harr}3"
        after_targets = ["ER_CPR_STOP_DATE", "ER_ANY_ROSC_DATE", "ER_SUS_ROSC_DATE",
                         "ER_DIE_DATE", "ER_OUT_DATE", "HOSP_AD_DATE"]
        seen_t = set()
        for fn in after_targets:
            if fn not in colmap:
                continue
            c = get_column_letter(colmap[fn])
            if c in seen_t:            # 합쳐진 @DT칸(날짜/시/분 동일열) 중복 방지
                continue
            seen_t.add(c)
            # 사이트는 '이전(<)'만 차단, 같은 시각은 허용 → '<'(DOA 사망=내원시각 오탐 방지).
            f = f"AND(ISNUMBER(${c}3),ISNUMBER({arr}),${c}3<{arr})"
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[f], fill=WARN_FILL))
            add_prompt(ws, c, n_rows, "⚠ 등록병원 응급실 내원시각(환자목록)보다 이후여야 "
                       "합니다. 이전이면 빨간색으로 표시됩니다.", "일시 규칙")

        # 병원 퇴원/사망일시는 '생존입원(HOSP_AD)' 이후여야(사이트는 내원 아닌 입원 기준).
        if "HOSP_AD_DATE" in colmap:
            hadc = f"${get_column_letter(colmap['HOSP_AD_DATE'])}3"
            for fn, lbl in [("HOSP_OUT_DATE", "병원퇴원일시(생존퇴원)"),
                            ("HOSP_DIE_DATE", "병원사망일시(사망퇴원)")]:
                _dt_order(ws, colmap, n_rows, fn, hadc, "<",
                          f"⚠ {lbl}는 생존입원 일시 이후여야 합니다. 이전이면 빨간색.")
        # 심폐소생술 중단은 내원 후 360분 이내(타병원경유는 상한 없음).
        if "ER_CPR_STOP_DATE" in colmap and "ER_LOCATION" in colmap:
            sc = get_column_letter(colmap["ER_CPR_STOP_DATE"])
            lc = get_column_letter(colmap["ER_LOCATION"])
            sS, eS = dt_serial(f"${sc}3"), dt_serial(arr)
            f = (f'AND(ISNUMBER(${sc}3),ISNUMBER({arr}),'
                 f'{sS}>{eS}+360/1440,${lc}3<>"타병원경유")')
            ws.conditional_formatting.add(
                f"{sc}3:{sc}{last}", FormulaRule(formula=[f], fill=WARN_FILL))
            add_prompt(ws, sc, n_rows, "⚠ 심폐소생술 중단은 내원시각 이후 360분 이내여야 "
                       "합니다(타병원경유 제외). 초과 시 빨간색.", "일시 규칙")


def apply_grey_community(ws, colmap, n_rows, patient_colmap=None):
    """지역사회: 심정지 목격여부(WITNESS)≠목격이면 발생(목격)추정시각 칸 회색.
    (사이트 JS는 비목격도 활성화하나, 사용자 요청상 '목격'만 활성.)"""
    if "WITNESS" not in colmap:
        return
    last = 3 + n_rows
    w = get_column_letter(colmap["WITNESS"])
    seen_cf = set()
    for fn in ["ONSET_DATE", "ONSET_CLOCK_HOUR", "ONSET_CLOCK_MIN", "ONSET_DATE_UK"]:
        if fn not in colmap:
            continue
        c = get_column_letter(colmap[fn])
        if c in seen_cf:  # 합쳐진 일시칸(날짜/시/분 동일칸) 중복 방지
            continue
        seen_cf.add(c)
        ws.conditional_formatting.add(
            f"{c}3:{c}{last}",
            FormulaRule(formula=[f'${w}3<>"목격"'], fill=GREY_FILL))

    # 목격자/발견자·발생장소의 '기타' 자유입력칸: 부모가 '기타'일 때만 흰색.
    for parent, etc in [("WIT_PERSON", "WIT_PERSON_ETC"),
                        ("ONSET_LOC", "ONSET_LOC_ETC")]:
        if parent in colmap and etc in colmap:
            pc = get_column_letter(colmap[parent])
            ec = get_column_letter(colmap[etc])
            ws.conditional_formatting.add(
                f"{ec}3:{ec}{last}",
                FormulaRule(formula=[f'${pc}3<>"기타"'], fill=GREY_FILL))

    # 목격자/발견자='근무중구급대원'이면 일반인 CPR/AED는 해당없음 → 회색.
    if "WIT_PERSON" in colmap:
        wp = get_column_letter(colmap["WIT_PERSON"])
        for fn in ("BYCPR", "BYDEFIB", "BYDEFIB_Y"):
            if fn in colmap:
                c = get_column_letter(colmap[fn])
                ws.conditional_formatting.add(
                    f"{c}3:{c}{last}",
                    FormulaRule(formula=[f'${wp}3="근무중구급대원"'], fill=GREY_FILL))

    # 목격(발생)추정시각은 내원시각의 '이전 24시간 이내'여야. 벗어나면 빨강.
    # 12자리 숫자를 실제 날짜시간으로 파싱해 [내원-1일, 내원] 밖이면 경고(월/년 경계 정확).
    if patient_colmap and "ER_DATE" in patient_colmap and "ONSET_DATE" in colmap and colmap:
        oc = get_column_letter(colmap["ONSET_DATE"])
        av = get_column_letter(patient_colmap["ER_DATE"])
        harr = add_xref_helper(ws, "환자목록", av, n_rows, max(colmap.values()) + 1)
        wd = dt_serial(f"${oc}3")        # 목격 시각(시리얼)
        ad = dt_serial(f"${harr}3")      # 내원 시각(시리얼, 헬퍼열)
        f = (f"AND(ISNUMBER(${oc}3),ISNUMBER(${harr}3),"
             f"OR({wd}<{ad}-1,{wd}>{ad}))")
        ws.conditional_formatting.add(
            f"{oc}3:{oc}{last}", FormulaRule(formula=[f], fill=WARN_FILL))
        add_prompt(ws, oc, n_rows, "⚠ 내원시각의 '이전 24시간 이내'여야 합니다. "
                   "24시간을 초과하거나 내원시각 이후이면 빨간색으로 표시됩니다.", "일시 규칙")


def apply_grey_prevent(ws, colmap, n_rows):
    """예방역학: 질환별 진단→치료여부→방법 연쇄 회색.
    진단(예/아니오/미상)≠예 → 치료여부 회색, 치료여부≠예 → 방법(@GROUP) 회색.
    (방법 그룹칸은 colmap 에 첫 멤버명 HTNTX1/DMTX1/DYSLTX1 로 기록됨.)"""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    # (진단, 치료여부, 방법그룹 첫멤버)
    for dx, tx, method in [("HTN", "HTNTX", "HTNTX1"),
                           ("DM", "DMTX", "DMTX1"),
                           ("DYSLIPID", "DYSLTX", "DYSLTX1")]:
        cdx, ctx, cm = L(dx), L(tx), L(method)
        if ctx and cdx:  # 치료여부: 진단≠예면 회색
            ws.conditional_formatting.add(
                f"{ctx}3:{ctx}{last}",
                FormulaRule(formula=[f'${cdx}3<>"예"'], fill=GREY_FILL))
        if cm and ctx:  # 방법: 치료여부≠예면 회색
            ws.conditional_formatting.add(
                f"{cm}3:{cm}{last}",
                FormulaRule(formula=[f'${ctx}3<>"예"'], fill=GREY_FILL))


def apply_grey_relief(ws, colmap, n_rows, patient_colmap=None):
    """구급단계: 부모 라디오 선택에 따라 종속 일시/상세/용량칸 회색 안내.
    - 일시칸(제세동/기계식압박/자발순환): 부모가 시행/적용/회복일 때만 흰색
    - 심정지리듬·기타, 기도확보 방법(그룹), 약물 상세(그룹): 부모 활성값일 때만 흰색
    - 약물별 총 용량: '약물 상세' 그룹칸에 해당 약이 선택됐을 때만 흰색(사이트와 동일)."""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    def rule(target, formula):
        c = L(target)
        if c:
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    # (부모필드, 흰색조건값, 대상칸) — 대상칸이 부모 활성값과 다르면 회색
    for parent, white, target in [
        ("PRE_DEFIB", "시행", "PRE_DEFIB_DATE"),         # 제세동 첫 시행일시
        ("PRE_MECHCPR", "적용", "PRE_MECHCPR_DATE"),     # 기계식압박 적용일시(신규)
        ("PRE_ROSC", "회복", "PRE_ROSC_DATE"),           # 자발순환 회복일시
        ("PRE_ECG", "심정지리듬", "PRE_ECG_RH"),          # 심정지리듬 종류
        ("PRE_ECG_RH", "기타", "PRE_ECG_RH_ETC"),        # 심정지리듬(기타) 자유입력
        ("FREE_AIRWAY", "확보함", "FREE_AIRWAY_OPA"),     # 기도확보 방법 그룹칸
        ("DRUG_USE", "사용함", "DRUG_USE_EPI"),           # 약물 상세 그룹칸
    ]:
        p = L(parent)
        if p:
            rule(target, f'${p}3<>"{white}"')

    # 약물별 총 용량: '약물 상세' 그룹칸(DRUG_USE_EPI 컬럼)에 해당 약 이름이 있어야 흰색.
    g = L("DRUG_USE_EPI")
    if g:
        for dose, drug in [("PRE_DRUG_EPI_TOT", "에피네프린"),
                           ("PRE_DRUG_AMIO_TOT", "아미오다론"),
                           ("PRE_DRUG_VASO_TOT", "바소프레신")]:
            rule(dose, f'NOT(ISNUMBER(SEARCH("{drug}",${g}3)))')

    # 제세동≠'시행'이면 첫 시행일시 칸 입력 자체를 차단(회색 안내 + 하드 잠금).
    # @DT 칸에 이미 붙은 프롬프트 DV 에 custom 규칙을 얹어 칸당 DV 1개 유지.
    pc, tc = L("PRE_DEFIB"), L("PRE_DEFIB_DATE")
    if pc and tc:
        rng = f"{tc}3:{tc}{last}"
        for dv in ws.data_validations.dataValidation:
            if str(dv.sqref) == rng:
                dv.type = "custom"
                dv.formula1 = f'${pc}3="시행"'
                dv.allow_blank = True         # 미시행/미상 → 빈칸 허용
                dv.showErrorMessage = True
                dv.errorTitle = "입력 불가"
                dv.error = '구급대 제세동이 "시행"일 때만 입력할 수 있습니다.'
                break

    # ── 재난번호 12자리 + 일시 순서/상한(빨강) ──────────────────────────
    if "DISASTER_NUM" in colmap:
        dcol = colmap["DISASTER_NUM"]
        dc = get_column_letter(dcol)
        for r in range(4, 4 + n_rows):          # 텍스트 서식: 앞자리 0 보존, LEN 정확
            ws.cell(row=r, column=dcol).number_format = "@"
        f = f'AND(${dc}4<>"",${dc}4<>"미상",LEN(${dc}4)<>12)'
        ws.conditional_formatting.add(
            f"{dc}4:{dc}{last}", FormulaRule(formula=[f], fill=WARN_FILL))
        add_prompt(ws, dc, n_rows, "⚠ 출동재난번호는 12자리여야 합니다('미상' 제외).", "형식 규칙")

    def cell(fn):
        return f"${get_column_letter(colmap[fn])}3" if fn in colmap else None

    scene = cell("SCENE_DATE")
    if cell("CALL_DATE"):
        _dt_window(ws, colmap, n_rows, "SCENE_DATE", cell("CALL_DATE"), 0, 600,
                   "⚠ 현장도착은 출동요청 이후 600분 이내여야 합니다.")
    if scene:
        _dt_window(ws, colmap, n_rows, "DEPARTURE_DATE", scene, 0, 600,
                   "⚠ 현장출발은 현장도착 이후 600분 이내여야 합니다.")
        _dt_window(ws, colmap, n_rows, "PRE_DEFIB_DATE", scene, 0, 600,
                   "⚠ 제세동은 현장도착 이후 600분 이내여야 합니다.")
    er = _er_helper_cell(ws, colmap, patient_colmap, n_rows)
    if er:
        _dt_window(ws, colmap, n_rows, "DEPARTURE_DATE", er, -600, 0,
                   "⚠ 현장출발은 내원시각 이전 600분 이내여야 합니다.")
        _dt_order(ws, colmap, n_rows, "PRE_DEFIB_DATE", er, ">",
                  "⚠ 제세동은 내원시각 이전이어야 합니다. 이후이면 빨간색.")
        _dt_order(ws, colmap, n_rows, "PRE_ROSC_DATE", er, ">",
                  "⚠ 병원전 자발순환회복은 내원시각 이전이어야 합니다. 이후이면 빨간색.")


def apply_grey_in_hosp(ws, colmap, n_rows, common_colmap=None, patient_colmap=None):
    """병원단계 조건부 회색 안내.
    #1 타병원 자발순환회복(PRE_ROSC): 공통영역 내원경로=타병원경유 일 때만 흰색(교차시트).
    #2 회복시각: PRE_ROSC=회복 / #3 심정지리듬: HOSP_ECG=심정지리듬
    #4 에피네프린 두 칸: 미상/미사용(EPINE 그룹) 선택시 회색
    #6/#7/#9 기타 자유입력칸: 부모=기타 일 때만 흰색."""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    def rule(target, formula):
        c = L(target)
        if c:
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    # 공통영역 값은 교차시트라 CF가 못 읽음 → 숨김 헬퍼열에 미러링해 같은 시트로 참조.
    hloc = hc1 = hc2 = None
    if common_colmap and colmap:
        s = SHEET_NAMES["common"]
        nextc = max(colmap.values()) + 1
        if "ER_LOCATION" in common_colmap:
            hloc = add_xref_helper(ws, s, get_column_letter(common_colmap["ER_LOCATION"]),
                                   n_rows, nextc); nextc += 1
        if "ER_CPR1" in common_colmap:
            hc1 = add_xref_helper(ws, s, get_column_letter(common_colmap["ER_CPR1"]),
                                  n_rows, nextc); nextc += 1
        if "ER_CPR2" in common_colmap:
            hc2 = add_xref_helper(ws, s, get_column_letter(common_colmap["ER_CPR2"]),
                                  n_rows, nextc); nextc += 1
    # #1: 공통영역 내원경로=타병원경유 아니면 PRE_ROSC 회색
    if hloc:
        rule("PRE_ROSC", f'${hloc}3<>"타병원경유"')
    # #2 회복시각
    pr = L("PRE_ROSC")
    if pr:
        rule("PRE_ROSC_DATE", f'${pr}3<>"회복"')
    # #3 심정지리듬
    he = L("HOSP_ECG")
    if he:
        rule("HOSP_ECG_RH", f'${he}3<>"심정지리듬"')
    # #4 에피네프린 3칸 게이팅:
    #   (A) 공통영역 심폐1·심폐2 중 하나라도 '미시행'이면 CPR 안 함 → 3칸 전부 회색(비활성).
    #   (B) 활성 상태 상호배타: 용량 2칸 입력시 미상/미사용 회색 / 미상·미사용 선택시 용량 2칸 회색.
    eg, et, er = L("EPINE_UK"), L("EPINE_TOT"), L("EPINE_ROSC")
    cpr_off = []  # 심폐 미시행 판정(헬퍼열). 심폐1은 '미시행 - …'라 앞 3글자로 판정.
    if hc1 and hc2:
        cpr_off = [f'LEFT(${hc1}3,3)="미시행"', f'${hc2}3="미시행"']
    if eg and et and er:
        def grey_if(target, extra):
            conds = cpr_off + extra
            rule(target, conds[0] if len(conds) == 1 else f'OR({",".join(conds)})')
        grey_if("EPINE_TOT", [f'${eg}3<>""'])   # 미상/미사용 선택시 회색
        grey_if("EPINE_ROSC", [f'${eg}3<>""'])
        grey_if("EPINE_UK", [f'${et}3<>""', f'${er}3<>""'])  # 용량 입력시 회색
    # #5 기관삽관≠'등록병원 시행'이면 첫 Advanced Airway 종류 회색
    ei = L("ENDO_INTU")
    if ei:
        rule("F_AD_AIRWAYD", f'${ei}3<>"등록병원 시행"')
    # ECMO 시행(EX_CIRC_EXE)이 미시행/미상(또는 미입력)이면 시술 관련 칸 전부 비활성 → 회색.
    # 시작·완료시점 / Pump-on·off 시점 / 기계 작동 성공.
    ce = L("EX_CIRC_EXE")
    if ce:
        for t in ("EX_CIR_S_DATE", "EX_CIR_E_DATE", "ECMO_DATE",
                  "ECMO5_POFF_DATE", "EX_CIR_SUCC"):
            rule(t, f'OR(${ce}3="",${ce}3="미시행",${ce}3="미상")')

    # #6/#7/#9 + Steroid 종류(기타): 부모=기타 일 때만 흰색
    for parent, etc in [("F_AD_AIRWAYD", "F_AD_AIRWAY_ETC"),
                        ("F_IV_KIND", "F_IV_KIND_ETC"),
                        ("ROSC_12ECG", "ROSC_12ECG_ETC"),
                        ("STEROID_KIND", "STEROID_KIND_ETC")]:
        p = L(parent)
        if p and etc in colmap:
            rule(etc, f'${p}3<>"기타"')

    # 단위 드롭다운: 짝 값칸이 ND(미측정)/미상이면 단위가 불필요 → 회색 안내.
    for name in list(colmap):
        if name.endswith("_UNIT") and name[:-5] in colmap:
            vc = get_column_letter(colmap[name[:-5]])  # 짝 값칸(@VUK 통합칸)
            rule(name, f'OR(${vc}3="ND",${vc}3="미상")')

    # Steroid 투여량(값칸+단위칸): 종류=미시행/미선택이면 회색 (약물명·기타일 때만 흰색)
    sk = L("STEROID_KIND")
    if sk:
        for t in ("STEROID", "STEROID_UNIT"):
            rule(t, f'OR(${sk}3="",${sk}3="미시행")')

    # 타병원 자발순환회복(PRE_ROSC)은 내원시각 이전이어야(빨강).
    er = _er_helper_cell(ws, colmap, patient_colmap, n_rows)
    if er:
        _dt_order(ws, colmap, n_rows, "PRE_ROSC_DATE", er, ">",
                  "⚠ 타병원 자발순환회복 일시는 내원시각 이전이어야 합니다. 이후이면 빨간색.")


def apply_grey_alive_after(ws, colmap, n_rows, common_colmap=None, patient_colmap=None):
    """소생후단계: 시술·검사·목표체온·승압제 등 '개별' 종속 회색만.
    시트 전체 게이팅 없음 — 사이트가 소생후를 공통 CPR 결과로 통째로 막지 않음
    (new_case6.html 확인: 전체-disable JS 없고 ROSC 직후 활력징후·승압제·목표체온
    상위칸이 활성, 페이지에 심폐1/2 값 자체가 없어 클라이언트 게이팅 불가)."""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    def rule(target, formula):
        c = L(target)
        if c:
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    # (시트 전체 게이팅 제거 2026-07-06: 사이트가 소생후를 공통 CPR 결과로 통째로
    #  막지 않음 — 개별 종속 회색만 유지. common_colmap 은 이제 미사용.)

    # 시술류(등록/타병원 시행이면 흰색): 미시행·미상·미선택이면 회색
    for parent, targets in [("REPER", ["REPER_TIME"]), ("CLOT", ["CLOT_DATE"]),
                            ("ANGIO", ["ANGIO_DATE"]), ("CA", ["CA_DATE"]),
                            ("CAB", ["CAB_DATE"])]:
        p = L(parent)
        if p:
            for t in targets:
                rule(t, f'OR(${p}3="",${p}3="미시행",${p}3="미상")')

    # 시행함류: '시행함' 아니면 회색 (ECMO7/영상검사 일시·결과/신경검사 결과)
    exec_map = {
        "ECMO7": ["ECMO7_DATE"],
        "SSEP": ["SSEP_DATE1", "SSEP_RESULT_CMMT", "SSEP_RESULT_FILE"],
        "EEG": ["EEG_DATE1", "EEG_RESULT_CMMT", "EEG_RESULT_FILE"],
        "BCT": ["BCT_DATE", "BCT_RESULT_CMMT", "BCT_RESULT_FILE"],
        "BMRI": ["BMRI_DATE", "BMRI_RESULT_CMMT", "BMRI_RESULT_FILE"],
        "GCS": ["GCS_E_RESULT", "GCS_V_RESULT", "GCS_M_RESULT"],
        "PLR": ["PLR_RESULT"], "CR": ["CR_RESULT"], "SR": ["SR_RESULT"],
        "GCS72": ["GCS72_E_RESULT", "GCS72_V_RESULT", "GCS72_M_RESULT"],
        "PLR72": ["PLR72_RESULT"], "CR72": ["CR72_RESULT"], "SR72": ["SR72_RESULT"],
    }
    for parent, targets in exec_map.items():
        p = L(parent)
        if p:
            for t in targets:
                rule(t, f'${p}3<>"시행함"')

    # 목표체온조절(LAW_TEMP): 사이트는 이 칸들을 막지 않음(항상 활성). 저장검증(checkInput)은
    # '미시행이면 세부에 미시행 외 표기 불가'라는 값 모순만 검사 → 회색 대신 빨강으로.
    lt = L("LAW_TEMP")
    if lt:
        last_r = 3 + n_rows
        # 시작/재가온 일시: 목표체온조절≠시행이면 비워두는 칸 → 회색 비활성화 안내.
        for t in ["LAW_TEMP_DATE", "RE_TEMP_DATE"]:
            rule(t, f'${lt}3<>"시행"')
        # 시행방법 4종: 미시행이면 '미시행' 또는 빈칸만 허용 → 그 외 값이면 빨강(모순).
        for t in ["EX_LAW_TEMP1", "EX_LAW_TEMP2", "EX_LAW_TEMP3", "IN_LAW_TEMP"]:
            c = L(t)
            if c:
                ws.conditional_formatting.add(
                    f"{c}3:{c}{last_r}",
                    FormulaRule(formula=[f'AND(${lt}3="미시행",${c}3<>"",${c}3<>"미시행")'],
                                fill=WARN_FILL))
                add_prompt(ws, c, n_rows,
                           "⚠ 목표체온조절=미시행이면 이 칸은 '미시행' 또는 빈칸이어야 합니다.", "모순 규칙")
        # 목표온도·재가온속도: '미시행' 옵션이 없음 → 미시행이면 빈칸이어야(값 있으면 빨강).
        for t in ["LAW_TEMP_TARGET", "RE_TEMP_SPEED"]:
            c = L(t)
            if c:
                ws.conditional_formatting.add(
                    f"{c}3:{c}{last_r}",
                    FormulaRule(formula=[f'AND(${lt}3="미시행",${c}3<>"")'], fill=WARN_FILL))
                add_prompt(ws, c, n_rows,
                           "⚠ 목표체온조절=미시행이면 이 칸은 비어 있어야 합니다.", "모순 규칙")

    # 승압제(HYPER): 'ROSC 24시간 이내'일 때만 종류 그룹 활성(미상은 드롭다운에 포함됨)
    hy = L("HYPER")
    if hy:
        rule("ROSC_INOTR_TYPE1", f'${hy}3<>"ROSC 24시간 이내"')

    # 기타 자유입력: 부모=기타 아니면 회색
    for parent, etc in [("LAW_TEMP_TARGET", "LAW_TEMP_TARGET_ETC"),
                        ("RE_TEMP_SPEED", "RE_TEMP_SPEED_ETC")]:
        p = L(parent)
        if p and etc in colmap:
            rule(etc, f'${p}3<>"기타"')
    # 승압제 종류 그룹에 '기타' 있을 때만 ETC 흰색
    g = L("ROSC_INOTR_TYPE1")
    if g and "ROSC_INOTR_TYPE_ETC" in colmap:
        rule("ROSC_INOTR_TYPE_ETC", f'NOT(ISNUMBER(SEARCH("기타",${g}3)))')

    # 재관류/시술(혈전용해·조영·성형·우회) 시행일시는 내원시각 이후여야(빨강).
    er = _er_helper_cell(ws, colmap, patient_colmap, n_rows)
    if er:
        for fn, lbl in [("CLOT_DATE", "혈전용해제"), ("ANGIO_DATE", "심혈관조영술"),
                        ("CA_DATE", "관상동맥성형술"), ("CAB_DATE", "관상동맥우회술")]:
            _dt_order(ws, colmap, n_rows, fn, er, "<",
                      f"⚠ {lbl} 시행일시는 내원시각 이후여야 합니다. 이전이면 빨간색.")


def apply_grey_heart(ws, colmap, n_rows, patient_colmap=None):
    """심장검사(항상 활성): 심장효소 ND→단위·일시 회색, 심초음파/CAG 미시행→하위 회색."""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    def rule(target, formula):
        c = L(target)
        if c:
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    # 심장효소: 값=ND/미상이면 그 항목의 단위칸·시행일시칸 회색
    for e in ["CKMB", "CKMB_PEAK", "TROP_I", "TROP_I_PEAK", "TROP_T", "TROP_T_PEAK",
              "BNP", "NTPRO_BNP"]:
        v = L(e)
        if v:
            for sfx in ["_UNIT", "_DATE"]:  # BNP류는 _DATE 없음 → colmap 없으면 skip
                rule(e + sfx, f'OR(${v}3="ND",${v}3="미상")')

    # 심초음파 시행함 아니면 국소벽장애(RWMA) 회색
    echo = L("ECHO_DONE1")
    if echo:
        rule("RWMA1", f'${echo}3<>"시행함"')

    # CAG 3시기: 미시행→결과(트리 드롭다운) 회색, 결과≠기타→기타 자유입력칸 회색
    for i in ("1", "2", "3"):
        done = L("CAG_DONE" + i)
        if done:
            rule("CAG_RE" + i, f'${done}3<>"시행함"')
        re = L("CAG_RE" + i)
        if re and ("CAG_RE_ETC" + i) in colmap:
            rule("CAG_RE_ETC" + i, f'${re}3<>"기타"')

    # 심장효소 시각(빨강): 내원시 검사는 내원 이후, peak 은 내원시 검사 이후.
    er = _er_helper_cell(ws, colmap, patient_colmap, n_rows)
    if er:
        for fn, lbl in [("CKMB_DATE", "내원시 CKMB"), ("TROP_I_DATE", "내원시 Troponin I"),
                        ("TROP_T_DATE", "내원시 Troponin T")]:
            _dt_order(ws, colmap, n_rows, fn, er, "<",
                      f"⚠ {lbl} 시행일시는 내원시각 이후여야 합니다. 이전이면 빨간색.")
    for peak, base, lbl in [("CKMB_PEAK_DATE", "CKMB_DATE", "CKMB peak"),
                            ("TROP_I_PEAK_DATE", "TROP_I_DATE", "Troponin I peak"),
                            ("TROP_T_PEAK_DATE", "TROP_T_DATE", "Troponin T peak")]:
        if base in colmap:
            bc = f"${get_column_letter(colmap[base])}3"
            _dt_order(ws, colmap, n_rows, peak, bc, "<",
                      f"⚠ {lbl} 시행일시는 내원시 검사 이후여야 합니다. 이전이면 빨간색.")


def apply_grey_y_child(ws, colmap, n_rows):
    """소아소생술(항상 활성, 게이팅 없음): 과거력=기타만 기타칸 활성,
    병원 퇴원시 PCPC=사망/미상이면 추적관찰(6·12개월) 회색,
    12개월 후 생존=생존/사망일 때만 12개월 후 PCPC 활성. 사망일시는 항상 활성."""
    last = 3 + n_rows

    def L(name):
        return get_column_letter(colmap[name]) if name in colmap else None

    def rule(target, formula):
        c = L(target)
        if c:
            ws.conditional_formatting.add(
                f"{c}3:{c}{last}", FormulaRule(formula=[formula], fill=GREY_FILL))

    mh = L("MEDI_HIST")
    if mh:
        rule("MEDI_HIST_ETC", f'${mh}3<>"기타"')

    # 병원 퇴원시 PCPC=PCPC6(사망)/미상(또는 미입력)이면 6·12개월 추적관찰 3칸 회색
    out = L("Y_OUT_PCPC")
    if out:
        dead = f'OR(${out}3="PCPC 6 (사망)",${out}3="미상",${out}3="")'
        for t in ("Y_6M_PCPC", "Y_12M_LIVE", "Y_12M_PCPC"):
            rule(t, dead)

    live = L("Y_12M_LIVE")
    if live:
        rule("Y_12M_PCPC", f'NOT(OR(${live}3="생존",${live}3="사망"))')


def apply_field_notes(ws, colmap):
    """FIELD_NOTES(필드명→설명글)를 헤더(1행) 셀에 Excel 메모로 부착.
    합쳐진 칸(@DT/@VUK/@TREE/@GROUP)은 구성 필드명 아무거로나 키하면 같은 열을 찾는다.
    봇은 셀 값만 읽으므로 메모는 크롤링에 영향 없음. 반환: 이 시트에서 찾은 필드명 집합."""
    hit = set()
    for name, text in FIELD_NOTES.items():
        col = colmap.get(name)
        if not col:
            continue
        hit.add(name)  # 열은 찾음(미매칭 경고 제외용)
        cell = ws.cell(row=1, column=col)
        if text and cell.comment is None:  # 같은 열에 이미 붙었으면 유지
            cm = Comment(str(text), "KOCARC")
            cm.width = 340
            cm.height = max(90, 16 * (str(text).count("\n") + 3))
            cell.comment = cm
    return hit


def _write_header(ws, col, label, name_text):
    """1행=라벨, 2행=필드명(또는 그룹마커) 헤더 작성."""
    c1 = ws.cell(row=1, column=col, value=label)
    c1.fill = HEADER_FILL
    c1.font = WHITE_BOLD
    c1.alignment = CENTER
    c1.border = BORDER
    c2 = ws.cell(row=2, column=col, value=name_text)
    c2.fill = NAME_FILL
    c2.font = NAME_FONT
    c2.alignment = CENTER
    c2.border = BORDER
    # 한글/CJK 는 2칸 폭이라 시각적 길이로 계산(헤더는 wrap 되므로 상한을 둔다).
    vis = sum(2 if ord(ch) >= 0x1100 else 1 for ch in str(label))
    ws.column_dimensions[get_column_letter(col)].width = max(12, min(32, round(vis * 0.7) + 3))


def write_field_header(ws, col, f, label=None):
    """1행=라벨, 2행=필드명 헤더 작성."""
    if label is None:
        label = display_label(f)
    _write_header(ws, col, label, f["name"])


def add_key_columns(ws, label_text):
    c1 = ws.cell(row=1, column=1, value=label_text)
    c1.fill = KEY_FILL
    c1.font = Font(bold=True, size=10)
    c1.alignment = CENTER
    c1.border = BORDER
    c2 = ws.cell(row=2, column=1, value="__KEY__")
    c2.fill = NAME_FILL
    c2.font = NAME_FONT
    c2.alignment = CENTER
    c2.border = BORDER
    ws.column_dimensions["A"].width = 10


def build_area_sheet(wb, area, sheet_title, n_rows=30, list_ctx=None,
                     common_colmap=None, patient_colmap=None):
    ws = wb.create_sheet(sheet_title)
    add_key_columns(ws, "환자키")
    count, colmap = write_columns(ws, area["fields"], n_rows, list_ctx)
    if area["key"] == "common":
        apply_grey_rules(ws, colmap, n_rows, patient_colmap)  # 종속 회색 + 시각순서 경고
    elif area["key"] == "prevent":
        apply_grey_prevent(ws, colmap, n_rows)
    elif area["key"] == "community":
        apply_grey_community(ws, colmap, n_rows, patient_colmap)
    elif area["key"] == "relief":
        apply_grey_relief(ws, colmap, n_rows, patient_colmap)
    elif area["key"] == "in_hosp":
        apply_grey_in_hosp(ws, colmap, n_rows, common_colmap, patient_colmap)
    elif area["key"] == "alive_after":
        apply_grey_alive_after(ws, colmap, n_rows, common_colmap, patient_colmap)
    elif area["key"] == "heart":
        apply_grey_heart(ws, colmap, n_rows, patient_colmap)
    elif area["key"] == "y_child":
        apply_grey_y_child(ws, colmap, n_rows)
    return count, colmap


def build_codebook(wb, schema):
    ws = wb.create_sheet("코드북")
    heads = ["영역", "질문(라벨)", "필드명", "유형", "허용값 (코드 = 의미)"]
    for j, h in enumerate(heads, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HEADER_FILL
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = BORDER
    widths = [14, 34, 20, 10, 70]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    def put_row(r, row):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=(j in (2, 5)))

    r = 2
    for area in schema["areas"]:
        for spec in grouped_columns(area["fields"]):
            if spec[0] == "group":
                _, glabel, members = spec
                labels = [(m["choices"][0].get("label") or "").strip() for m in members]
                desc = "택1 또는 복수(쉼표): " + " | ".join(labels)
                gdisp = with_context(glabel, members[0].get("context"))
                put_row(r, [area["title"], gdisp, "(드롭다운)", "드롭다운", desc])
                r += 1
                continue
            f = spec[1]
            if f.get("tree"):
                byname = {x["name"]: x for x in area["fields"]}
                leaves, _ = tree_leaves(byname, f)
                desc = "택1(부모-자식 통합): " + " | ".join(leaves)
                put_row(r, [area["title"], display_label(f), f["name"],
                            "조건부", desc])
                r += 1
                continue
            pairs = []
            if f["type"] in ("radio", "checkbox"):
                pairs = [(c["value"], c["label"]) for c in f["choices"]]
            elif f["type"] == "select":
                pairs = [(o["value"], o["label"]) for o in f["options"]
                         if o["value"] not in ("", "XXXX", "XX", "--", "XX:-", "--:-")]
            if not pairs:
                continue  # 자유입력(텍스트/날짜/숫자)은 코드북에 안 실음
            if len(pairs) > 40:
                desc = f"(목록 {len(pairs)}개: 연/시각 등 숫자 선택)"
            else:
                desc = " | ".join(f"{v} = {t}" for v, t in pairs)
            put_row(r, [area["title"], display_label(f), f["name"], f["type"], desc])
            r += 1
    ws.freeze_panes = "A2"


def build_usage(wb):
    ws = wb.create_sheet("사용법")
    lines = [
        "KOCARC 환자 자동등록 입력 양식",
        "",
        "■ 작성 방법",
        "1) [환자목록] 시트에 환자를 한 줄씩 적습니다. (A열 환자키 = 1,2,3 ... 미리 채워져 있음)",
        "2) 각 영역 시트(공통영역, 예방역학 ...)에서 같은 '환자키' 행에 값을 채웁니다.",
        "3) 값은 드롭다운에서 고르거나 [코드북]을 참고하세요. 코드(예: M)·의미(예: Male) 둘 다 됩니다.",
        "4) 해당 없는 칸은 비워두면 됩니다. (빈칸은 입력하지 않음)",
        "",
        "■ 시트 구조 (중요)",
        " - 1행 = 질문(한글 라벨). 셀 모서리에 빨간 삼각형이 있으면 마우스를 올려 도움말을 보세요.",
        " - 2행 = 필드명(프로그램이 읽는 줄) → 절대 수정하지 마세요.",
        " - 3행 = 회색 예시(입력 형식 참고용). 여기에는 입력하지 마세요.",
        " - 4행부터 = 실제 환자 데이터.",
        "",
        "■ 날짜 / 시각 (한 칸에 적습니다)",
        " - 시각 칸: '2026-06-15 09:30' 또는 숫자 12자리 '202606150930' → 자동으로 형식이 맞춰집니다.",
        "   (시각을 모르면 날짜만 적어도 됩니다.)",
        " - 생년월일 칸: '1970-05-15' 또는 숫자 8자리 '19700515' (한 칸).",
        "",
        "■ 드롭다운 / 색 표시 (회색·빨강)",
        " - 목록이 있는 칸은 목록에서 고르세요. 일부 칸은 쉼표(,)로 여러 개 선택할 수 있습니다(셀 클릭 시 안내).",
        " - 회색 칸: 앞 칸 선택에 따라 필요 없는 칸입니다. (안내일 뿐이며 입력이 막히진 않습니다.)",
        " - 빨강 칸: 입력한 값이 규칙에 맞지 않을 때 표시됩니다(필수항목 빈칸, 일시 순서·범위 오류 등). 경고일 뿐 입력·저장은 됩니다.",
        "",
        "■ 검사결과 파일 (SSEP·EEG·뇌CT·뇌MRI 등)",
        " - 결과 '파일 첨부'는 프로그램이 대신 올리지 못합니다. 필요하면 등록 후 사이트에서 직접 업로드하세요.",
        " - 엑셀에는 검사결과 '소견(텍스트)'만 적으면 됩니다. (파일 칸은 양식에 없습니다.)",
        "",
        "■ 주의",
        " - 비밀번호는 이 파일에 적지 마세요.",
        " - 자동 등록은 실제 연구 DB에 저장됩니다. 처음에는 프로그램 '특정 환자키만'에 1 을 넣어 1명만 시험하세요.",
        " - 데이터를 채운 뒤에는 '양식 다시 만들기'를 누르지 마세요. (빈 양식으로 덮어써져 입력이 사라집니다.)",
    ]
    for i, t in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=t)
        if i == 1:
            c.font = Font(bold=True, size=14, color="2E6B4F")
        elif t.startswith("■"):
            c.font = Font(bold=True, size=11, color="2E6B4F")
    ws.column_dimensions["A"].width = 95


def build(out_path=None, schema_path=None, log=print):
    out_path = out_path or OUT
    schema = load_schema(schema_path)
    areas = {a["key"]: a for a in schema["areas"]}

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    build_usage(wb)
    list_ctx = ListSheet(wb)  # 콤마/따옴표 든 라벨용 숨김 목록 시트

    # 환자목록 = 환자등록(시작) 영역
    pa = areas.get("patient_add")
    ws = wb.create_sheet("환자목록")
    add_key_columns(ws, "환자키")
    pa_count, pa_colmap = write_columns(ws, pa["fields"], 30, list_ctx)
    apply_required_rules(ws, pa_colmap, 30, log=log)   # 필수항목 빈칸 → 빨강

    # 나머지 영역 시트 (입력칸이 없는 영역은 시트 생성 안 함)
    # common 을 먼저 만들어 colmap 을 보관 → in_hosp 교차시트 회색규칙에서 참조.
    counts, colmaps = {}, {}
    for key in ["common", "prevent", "community", "relief", "in_hosp",
                "alive_after", "heart", "y_child", "comment"]:
        if key not in areas:
            continue
        if not any(f["user_entry"] for f in areas[key]["fields"]):
            continue
        counts[key], colmaps[key] = build_area_sheet(
            wb, areas[key], SHEET_NAMES[key], list_ctx=list_ctx,
            common_colmap=colmaps.get("common"), patient_colmap=pa_colmap)

    build_codebook(wb, schema)

    # 항목별 도움말 메모 부착 (환자목록 + 각 영역 시트)
    if FIELD_NOTES:
        matched = apply_field_notes(ws, pa_colmap)
        for key, cmap in colmaps.items():
            matched |= apply_field_notes(wb[SHEET_NAMES[key]], cmap)
        unmatched = [k for k in FIELD_NOTES if k not in matched]
        if unmatched:
            log(f"⚠ 메모 미매칭 필드명 {len(unmatched)}개(오타/제외항목?): {unmatched}")

    wb.save(out_path)
    log(f"엑셀 양식 저장: {out_path}")
    total = pa_count + sum(counts.values())
    log(f"전체 입력 칸: {total}개 (환자목록 {pa_count} + 영역들)")
    return out_path


def main():
    build()


if __name__ == "__main__":
    main()
