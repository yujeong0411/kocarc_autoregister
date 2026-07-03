# -*- coding: utf-8 -*-
"""
KOCARC eCRF 자동 등록 봇.

흐름:
  1) 로그인
  2) 엑셀 [환자목록] 행마다 환자 생성 -> 새 PAT_ID 획득
  3) 각 영역 시트의 같은 환자키 행 값으로 영역 페이지를 채우고 저장
  4) 진행상황을 progress.csv 에 기록 (중단 후 재실행 시 이어서)

안전장치:
  - 실제 연구 DB에 생성·저장됨. 처음에는 only_keys 로 1명만 시험 후 전체 실행 권장.
  - 이미 done 인 환자는 progress.csv 로 건너뜀(중복 등록 방지).

주의: 비밀번호는 config.ini(본인 PC) 또는 실행 시 입력. 코드/엑셀에 적지 말 것.
"""
import os
import re
import sys
import csv
import time
import json
import getpass
import configparser
from datetime import datetime

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        NoAlertPresentException, UnexpectedAlertPresentException,
        NoSuchElementException, TimeoutException,
    )
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))


def resource_path(name):
    """읽기전용 동봉 파일(schema.json 등). PyInstaller 번들 대응."""
    base = getattr(sys, "_MEIPASS", BASE)
    return os.path.join(base, name)


def app_dir():
    """쓰기 가능한 작업 폴더(progress.csv 등). exe 옆 또는 소스 폴더."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return BASE


# ---- 영역 정의: (스키마키, 엑셀시트명, 영역 add 페이지 경로) ----
AREAS = [
    ("common",      "공통영역",   "/kocarc/patient/patient_up.asp"),
    ("prevent",     "예방역학",   "/kocarc/patient/prevent/prevent_add.asp"),
    ("community",   "지역사회",   "/kocarc/patient/community/community_add.asp"),
    ("relief",      "구급단계",   "/kocarc/patient/relief/relief_add.asp"),
    ("in_hosp",     "병원단계",   "/kocarc/patient/in_hosp/in_hosp_add.asp"),
    ("alive_after", "소생후단계", "/kocarc/patient/alive_after/alive_after_add.asp"),
    ("heart",       "심장검사",   "/kocarc/patient/heart/heart_add.asp"),
    ("y_child",     "소아소생술", "/kocarc/patient/y_child/y_child_add.asp"),
    ("comment",     "CommentLog", "/kocarc/patient/patient_comment_add.asp"),
]
PATIENT_ADD_PATH = "/kocarc/patient/patient_add.asp"


_LOG_FN = None


def set_log(fn):
    """로그 출력을 다른 곳(GUI 창 등)으로 보낼 때 콜백 등록."""
    global _LOG_FN
    _LOG_FN = fn


def log(msg):
    line = f"[{datetime.now():%H:%M:%S}] {msg}"
    if _LOG_FN:
        try:
            _LOG_FN(line)
            return
        except Exception:
            pass
    print(line, flush=True)


# =====================================================================
# 설정
# =====================================================================
def load_config():
    cfg = configparser.ConfigParser()
    path = os.path.join(app_dir(), "config.ini")
    if not os.path.exists(path):
        log("config.ini 가 없습니다. config.example.ini 를 복사해 만드세요.")
        sys.exit(1)
    cfg.read(path, encoding="utf-8")
    c = cfg["kocarc"]
    conf = {
        "base_url": c.get("base_url", "https://ecrf.kr").rstrip("/"),
        "login_url": c.get("login_url", "https://ecrf.kr/kocarc/"),
        "member_id": c.get("member_id", "kocarc_14"),
        "password": c.get("password", ""),
        "excel": c.get("excel", "KOCARC_입력양식.xlsx"),
        "headless": c.getboolean("headless", False),
        "areas": [a.strip() for a in c.get("areas", "all").split(",")],
        "pause": c.getfloat("pause", 0.3),
        "only_keys": [k.strip() for k in c.get("only_keys", "").split(",") if k.strip()],
    }
    if not conf["password"]:
        conf["password"] = getpass.getpass("KOCARC 비밀번호 입력: ")
    return conf


# 흔한 한글 표현 -> 폼 라벨에 들어있는 단어 (직접/라벨 매칭 실패 시 보조)
SYNONYMS = {
    "남": "male", "남자": "male", "남성": "male",
    "여": "female", "여자": "female", "여성": "female",
}


# =====================================================================
# 스키마 기반 값 변환 (사람이 적은 코드/의미 -> 폼 전송값)
# =====================================================================
class Resolver:
    def __init__(self, schema):
        self.fields = {}   # (area_key, name) -> field dict
        for a in schema["areas"]:
            for f in a["fields"]:
                self.fields[(a["key"], f["name"])] = f

    def field(self, area, name):
        return self.fields.get((area, name))

    def to_form_value(self, area, name, raw):
        """엑셀 값(raw) -> 실제 폼 값. 매칭 실패 시 원문 그대로 반환."""
        if raw is None:
            return None
        s = str(raw).strip()
        if s == "":
            return None
        f = self.field(area, name)
        if not f:
            return s
        pairs = []
        if f["type"] in ("radio", "checkbox"):
            pairs = [(c["value"], c["label"]) for c in f["choices"]]
        elif f["type"] == "select":
            pairs = [(o["value"], o["label"]) for o in f["options"]]
        if not pairs:
            return s  # 자유입력
        low = s.lower()
        for v, lab in pairs:
            if s == v:
                return v
        for v, lab in pairs:
            if low == (lab or "").lower():
                return v
        # 부분일치(앞부분)
        for v, lab in pairs:
            if lab and low == lab.lower().split()[0]:
                return v
        # 한글 동의어 보조 (예: 남자 -> male -> M)
        syn = SYNONYMS.get(s)
        if syn:
            for v, lab in pairs:
                if syn in (lab or "").lower():
                    return v
        return s  # 못 찾으면 원문 (로그로 경고)


# =====================================================================
# 엑셀 읽기
# =====================================================================
def read_sheet_rows(wb, sheet_name):
    """시트를 읽어 {환자키: {필드명: 값}} 로 반환. (1행=라벨, 2행=필드명, 3행~데이터)"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    names = {}
    for col in range(1, ws.max_column + 1):
        nm = ws.cell(row=2, column=col).value
        if nm:
            names[col] = str(nm).strip()
    out = {}
    for r in range(3, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if key is None or str(key).strip() == "":
            continue
        row = {}
        for col, nm in names.items():
            if nm in ("__KEY__",):
                continue
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                row[nm] = v
        if row:
            out[str(key).strip()] = row
    return out


def read_all(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    patients = read_sheet_rows(wb, "환자목록")     # {key: {field:val}}
    area_data = {}
    for key, sheet, _ in AREAS:
        area_data[key] = read_sheet_rows(wb, sheet)
    return patients, area_data


# =====================================================================
# Selenium 헬퍼
# =====================================================================
SET_VALUE_JS = r"""
var form = document.form1; if(!form) return 'NOFORM';
var name = arguments[0], val = arguments[1], kind = arguments[2];
var els = form.elements[name];
if(!els) return 'NOFIELD';
function fire(el){ ['input','change'].forEach(function(t){ try{ el.dispatchEvent(new Event(t,{bubbles:true})); }catch(e){} }); }
function enable(el){ try{ el.disabled=false; }catch(e){} }
// 사이트는 상위 항목(예: 병원진단=예) 클릭 전엔 하위칸을 disabled 로 둔다.
// 봇은 onclick 의존 로직을 안 타므로, 채우려는 칸은 먼저 활성화한다.
if(kind==='radio' || kind==='checkbox'){
   var list = els.length ? els : [els];
   // 단일 체크박스(부/모 등 value 없는 '해당시 체크' 칸): 값이 들어오면 무조건 체크
   if(kind==='checkbox' && list.length===1){
       enable(list[0]); list[0].checked=true; fire(list[0]); return 'OK';
   }
   var hit=false;
   for(var i=0;i<list.length;i++){
       if(String(list[i].value)===String(val)){ enable(list[i]); list[i].checked=true; fire(list[i]); hit=true; }
       else if(kind==='radio'){ list[i].checked=false; }
   }
   return hit ? 'OK' : 'NOVAL';
} else if(els.tagName==='SELECT'){
   enable(els);
   var ok=false;
   for(var i=0;i<els.options.length;i++){ if(String(els.options[i].value)===String(val)){ els.selectedIndex=i; ok=true; break; } }
   if(!ok){ els.value=val; }
   fire(els);
   return ok ? 'OK' : 'SETRAW';
} else {
   // text / textarea / hidden 등
   var t = els.length ? els[0] : els;
   enable(t);
   t.value = val; fire(t);
   return 'OK';
}
"""


def make_driver(conf):
    opts = webdriver.ChromeOptions()
    if conf["headless"]:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,1000")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(60)
    return driver


def dismiss_alert(driver):
    try:
        a = driver.switch_to.alert
        txt = a.text
        a.accept()
        return txt
    except NoAlertPresentException:
        return None


def login(driver, conf):
    log(f"로그인 페이지 이동: {conf['login_url']}")
    driver.get(conf["login_url"])
    WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.NAME, "MemberId")))
    driver.find_element(By.NAME, "MemberId").clear()
    driver.find_element(By.NAME, "MemberId").send_keys(conf["member_id"])
    driver.find_element(By.NAME, "MemberPassword").send_keys(conf["password"])
    driver.execute_script("document.signform.submit();")
    time.sleep(2)
    dismiss_alert(driver)
    # 로그인 성공 판정: 로그인 폼이 사라졌는지
    if driver.find_elements(By.NAME, "MemberPassword"):
        raise RuntimeError("로그인 실패 — 아이디/비밀번호를 확인하세요.")
    log("로그인 성공")


def find_pat_id(driver):
    """현재 페이지에서 새로 생성된 PAT_ID 추출."""
    import re
    url = driver.current_url
    m = re.search(r"PAT_ID=([0-9]{2}-[0-9]{4,6})", url)
    if m:
        return m.group(1)
    for el in driver.find_elements(By.NAME, "PAT_ID"):
        v = el.get_attribute("value")
        if v and re.match(r"[0-9]{2}-[0-9]{4,6}", v):
            return v
    m = re.search(r"\b[0-9]{2}-[0-9]{4,6}\b", driver.page_source)
    return m.group(0) if m else None


GROUP_PREFIX = "@GROUP:"


def fill_group(driver, resolver, area_key, group_name, raw, conf):
    """드롭다운 1칸으로 묶인 체크박스 그룹 처리.
    엑셀 칸 값(부/모/... 또는 쉼표로 복수)을 해당 체크박스 필드명으로 바꿔 체크한다.
    반환: (성공개수, 경고리스트)"""
    members = [m.strip() for m in group_name[len(GROUP_PREFIX):].split(",") if m.strip()]
    # 라벨/코드/필드명 -> 실제 필드명 매핑
    label_map = {}
    for mn in members:
        mf = resolver.field(area_key, mn)
        ch = (mf.get("choices") or [{}])[0] if mf else {}
        lab = (ch.get("label") or "").strip()
        if lab:
            label_map[lab.lower()] = mn
        val = str(ch.get("value") or "").strip()
        if val:
            label_map[val.lower()] = mn
        label_map[mn.lower()] = mn  # 필드명 직접 입력도 허용
    ok, warn = 0, []
    for tok in re.split(r"[,/]", str(raw)):
        t = tok.strip()
        if not t:
            continue
        target = label_map.get(t.lower())
        if not target:  # 첫 단어 부분일치 보조
            for lab, mn in label_map.items():
                if lab.split() and lab.split()[0] == t.lower():
                    target = mn
                    break
        if not target:
            warn.append(f"{t}(매칭실패)")
            continue
        try:
            res = driver.execute_script(SET_VALUE_JS, target, "1", "checkbox")
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        if res == "OK":
            ok += 1
        else:
            warn.append(f"{target}({res})")
        if conf["pause"]:
            time.sleep(conf["pause"] / 10.0)
    return ok, warn


# 조건부 라디오→라디오 트리 (build_template.TREE_PREFIX/TREE_SEP 와 동일해야 함)
TREE_PREFIX = "@TREE:"
TREE_SEP = " - "


def parse_tree_marker(marker):
    """@TREE:부모|부모값=자식;부모값=자식  →  (부모필드, {부모값: 자식필드})"""
    parent, _, mapstr = marker[len(TREE_PREFIX):].partition("|")
    tree = {}
    for kv in mapstr.split(";"):
        if "=" in kv:
            pv, cn = kv.split("=", 1)
            tree[pv.strip()] = cn.strip()
    return parent.strip(), tree


def tree_targets(resolver, area_key, marker, raw):
    """마커 + 엑셀값('부모 - 자식(- 손자)' 또는 '부모라벨') → [(필드, 폼값)].
    직속 자식은 마커에서, 더 깊은 단계는 각 자식필드의 schema tree 로 재귀 복원한다.
    순수 함수(브라우저 불필요) — 자체검증에서 재사용."""
    parent, tree = parse_tree_marker(marker)
    parts = [p.strip() for p in str(raw).split(TREE_SEP) if p.strip()]
    out, field, childmap = [], parent, tree
    for part in parts:
        val = resolver.to_form_value(area_key, field, part)
        out.append((field, val))
        child = childmap.get(str(val))
        if not child:
            break
        field = child
        # 다음 단계 자식맵: schema 의 자식필드 tree (없으면 종료)
        fdef = resolver.field(area_key, field) or {}
        childmap = fdef.get("tree") or {}
    return out


# 날짜+시+분 통합칸 (build_template.DT_PREFIX 와 동일해야 함)
DT_PREFIX = "@DT:"


def datetime_targets(raw):
    """'2026-06-15 05:02'(또는 구분자없이 202606150502, 날짜만 20260615) →
    (날짜, 시, 분) 문자열. 없으면 None. 시/분은 2자리 0채움. 날짜만이면 시/분 None."""
    s = str(raw).strip()
    date = hour = minute = None
    # 구분자 없는 순수 숫자: 8자리 YYYYMMDD(날짜만) 또는 12자리 YYYYMMDDHHMM
    g = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(?:(\d{2})(\d{2}))?", s)
    if g:
        date = f"{g.group(1)}-{g.group(2)}-{g.group(3)}"
        if g.group(4):
            hour, minute = g.group(4), g.group(5)
        return date, hour, minute
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        date = f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    t = re.search(r"(\d{1,2}):(\d{1,2})", s)
    if t:
        hour = f"{int(t.group(1)):02d}"
        minute = f"{int(t.group(2)):02d}"
    return date, hour, minute


def fill_datetime(driver, resolver, area_key, marker, raw, conf):
    """@DT 통합칸: '2026-06-15 05:02'를 날짜/시/분 필드로 분해해 입력.
    마커에 4번째 필드(_UK 미상 체크박스)가 있고 셀에 '미상'이 적혀 있으면
    날짜/시/분 대신 그 체크박스만 체크한다."""
    fields = marker[len(DT_PREFIX):].split(",")
    uk = fields[3] if len(fields) > 3 else None
    if uk and "미상" in str(raw):
        try:
            res = driver.execute_script(SET_VALUE_JS, uk, "1", "checkbox")
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        return (1, []) if res == "OK" else (0, [f"{uk}=미상({res})"])
    date, hour, minute = datetime_targets(raw)
    ok, warn = 0, []
    for fn, val in zip(fields[:3], (date, hour, minute)):
        if val is None:
            continue
        f = resolver.field(area_key, fn)
        kind = f["type"] if f else "text"
        try:
            res = driver.execute_script(SET_VALUE_JS, fn, str(val), kind)
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        if res == "OK":
            ok += 1
        else:
            warn.append(f"{fn}={val}({res})")
    return ok, warn


# 생년월일 년/월/일 통합칸 (build_template.DOB_PREFIX 와 동일해야 함)
DOB_PREFIX = "@DOB:"


def dob_targets(raw):
    """'1970-05-15'(또는 1970.5.9, 구분자없이 19700515) → (년4, 월2, 일2) 셀렉트
    value 문자열. 년은 옵션값이 '1970'이라 0채움 없음, 월/일은 2자리.
    구분자 없이 넣으면 월·일 구분이 안되므로 정확히 8자리만 허용. 형식 안 맞으면 None."""
    s = str(raw).strip()
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if not m:
        m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})", s)  # 구분자 없이 YYYYMMDD
    if not m:
        return None
    return (str(int(m.group(1))), f"{int(m.group(2)):02d}", f"{int(m.group(3)):02d}")


def fill_dob(driver, resolver, area_key, marker, raw, conf):
    """@DOB 통합칸: 'YYYY-MM-DD'를 년/월/일 셀렉트로 분해해 선택."""
    fields = marker[len(DOB_PREFIX):].split(",")  # [년, 월, 일]
    vals = dob_targets(raw)
    if vals is None:
        return 0, [f"{fields[0]}=생년월일형식({raw})"]
    ok, warn = 0, []
    for fn, val in zip(fields, vals):
        f = resolver.field(area_key, fn)
        kind = f["type"] if f else "select"
        try:
            res = driver.execute_script(SET_VALUE_JS, fn, val, kind)
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        if res == "OK":
            ok += 1
        else:
            warn.append(f"{fn}={val}({res})")
    return ok, warn


# 값칸 + 미상(_UK) 통합칸 (build_template.VUK_PREFIX 와 동일해야 함)
VUK_PREFIX = "@VUK:"


def fill_valuk(driver, resolver, area_key, marker, raw, conf):
    """@VUK 통합칸: 값칸에 키워드(미상/ND 등)를 적으면 플래그 체크박스를 누르고,
    아니면 값칸에 값 입력. 마커 형식: @VUK:값필드,체크박스,키워드 (키워드 생략시 '미상')."""
    parts = marker[len(VUK_PREFIX):].split(",")
    val_field, uk_field = parts[0], parts[1]
    keyword = parts[2] if len(parts) > 2 else "미상"
    if str(raw).strip().lower() == keyword.lower():
        try:
            res = driver.execute_script(SET_VALUE_JS, uk_field, "1", "checkbox")
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        return (1, []) if res == "OK" else (0, [f"{uk_field}={keyword}({res})"])
    formval = resolver.to_form_value(area_key, val_field, raw)
    if formval is None:
        return 0, []
    f = resolver.field(area_key, val_field)
    kind = f["type"] if f else "text"
    try:
        res = driver.execute_script(SET_VALUE_JS, val_field, str(formval), kind)
    except UnexpectedAlertPresentException:
        dismiss_alert(driver)
        res = "ALERT"
    return (1, []) if res == "OK" else (0, [f"{val_field}={raw}({res})"])


def fill_tree(driver, resolver, area_key, marker, raw, conf):
    """@TREE 통합칸: 부모 라디오 + (해당 시) 자식 라디오를 채운다.
    반환: (성공개수, 경고리스트)"""
    ok, warn = 0, []
    for field, val in tree_targets(resolver, area_key, marker, raw):
        if val is None:
            continue
        try:
            res = driver.execute_script(SET_VALUE_JS, field, str(val), "radio")
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        if res == "OK":
            ok += 1
        else:
            warn.append(f"{field}={val}({res})")
        if conf["pause"]:
            time.sleep(conf["pause"] / 10.0)
    return ok, warn


def fill_form1(driver, resolver, area_key, values, conf):
    """현재 페이지 form1 에 values 를 채운다. (저장은 별도)"""
    set_ok, warn = 0, []
    for name, raw in values.items():
        if name.startswith(DT_PREFIX):
            d_ok, d_warn = fill_datetime(driver, resolver, area_key, name, raw, conf)
            set_ok += d_ok
            warn.extend(d_warn)
            continue
        if name.startswith(DOB_PREFIX):
            b_ok, b_warn = fill_dob(driver, resolver, area_key, name, raw, conf)
            set_ok += b_ok
            warn.extend(b_warn)
            continue
        if name.startswith(VUK_PREFIX):
            v_ok, v_warn = fill_valuk(driver, resolver, area_key, name, raw, conf)
            set_ok += v_ok
            warn.extend(v_warn)
            continue
        if name.startswith(TREE_PREFIX):
            t_ok, t_warn = fill_tree(driver, resolver, area_key, name, raw, conf)
            set_ok += t_ok
            warn.extend(t_warn)
            continue
        if name.startswith(GROUP_PREFIX):
            g_ok, g_warn = fill_group(driver, resolver, area_key, name, raw, conf)
            set_ok += g_ok
            warn.extend(g_warn)
            continue
        f = resolver.field(area_key, name)
        kind = f["type"] if f else "text"
        formval = resolver.to_form_value(area_key, name, raw)
        if formval is None:
            continue
        try:
            res = driver.execute_script(SET_VALUE_JS, name, str(formval), kind)
        except UnexpectedAlertPresentException:
            dismiss_alert(driver)
            res = "ALERT"
        if res == "OK":
            set_ok += 1
        else:
            warn.append(f"{name}={raw}({res})")
        if conf["pause"]:
            time.sleep(conf["pause"] / 10.0)
    if warn:
        log(f"   ! 확인필요({area_key}): " + ", ".join(warn[:15]) +
            (" ..." if len(warn) > 15 else ""))
    return set_ok


MARK_COMPLETE_JS = r"""
var f = document.form1; if(!f) return 'NOFORM';
var e = f.elements['COMPLETE_YN']; if(!e) return 'NONE';
e = e.length ? e[0] : e;
e.checked = true;
try{ e.dispatchEvent(new Event('change',{bubbles:true})); }catch(x){}
return 'OK';
"""


def mark_complete(driver):
    """현재 영역 페이지의 '핵심변수 입력완료'(COMPLETE_YN) 체크박스를 체크."""
    try:
        res = driver.execute_script(MARK_COMPLETE_JS)
    except UnexpectedAlertPresentException:
        dismiss_alert(driver)
        res = "ALERT"
    if res == "OK":
        log("   입력완료(핵심변수) 체크")
    return res


def save_form1(driver, conf):
    try:
        driver.execute_script("document.form1.submit();")
    except UnexpectedAlertPresentException:
        pass
    time.sleep(1.0)
    txt = dismiss_alert(driver)
    if txt:
        log(f"   저장 후 알림: {txt}")
    time.sleep(0.5)
    return True


# =====================================================================
# 메인 처리
# =====================================================================
def create_patient(driver, conf, resolver, prow):
    url = conf["base_url"] + PATIENT_ADD_PATH
    driver.get(url)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "PAT_NM")))
    fill_form1(driver, resolver, "patient_add", prow, conf)
    driver.execute_script("document.form1.submit();")
    time.sleep(1.5)
    txt = dismiss_alert(driver)
    if txt:
        log(f"   등록 알림: {txt}")
    time.sleep(1.0)
    pat_id = find_pat_id(driver)
    return pat_id


def load_progress():
    path = os.path.join(app_dir(), "progress.csv")
    done = {}
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                done[row["key"]] = row
    return done


def append_progress(key, pat_id, status):
    path = os.path.join(app_dir(), "progress.csv")
    new = not os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["time", "key", "pat_id", "status"])
        w.writerow([datetime.now().isoformat(timespec="seconds"), key, pat_id or "", status])


def run_bot(conf, should_stop=None):
    """봇 본체. GUI/CLI 공통 진입점.
    conf: 설정 dict (password 포함)
    should_stop: 중지 요청 확인용 콜백(있으면 True 일 때 중단)
    """
    def stop():
        return bool(should_stop and should_stop())

    if not SELENIUM_OK:
        log("selenium 이 설치되어 있지 않습니다. (uv sync 또는 pip install selenium)")
        return
    schema_path = resource_path("schema.json")
    with open(schema_path, encoding="utf-8") as f:
        schema = json.load(f)
    resolver = Resolver(schema)

    excel = conf["excel"]
    excel_path = excel if os.path.isabs(excel) else os.path.join(app_dir(), excel)
    if not os.path.exists(excel_path):
        log(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return
    patients, area_data = read_all(excel_path)
    if not patients:
        log("환자목록 시트에 데이터가 없습니다.")
        return

    want_areas = [a for a in AREAS
                  if conf["areas"] == ["all"] or a[0] in conf["areas"]]
    keys = list(patients.keys())
    if conf["only_keys"]:
        keys = [k for k in keys if k in conf["only_keys"]]

    log(f"대상 환자 {len(keys)}명, 영역 {len(want_areas)}개 — 실제 생성·저장")
    progress = load_progress()

    driver = make_driver(conf)
    try:
        login(driver, conf)
        for key in keys:
            if stop():
                log("중지 요청 — 종료합니다.")
                break
            if key in progress and progress[key]["status"] == "done":
                log(f"환자키 {key}: 이미 완료 — 건너뜀")
                continue
            log(f"=== 환자키 {key} 처리 시작 ===")
            try:
                pat_id = create_patient(driver, conf, resolver, patients[key])
                if not pat_id:
                    log("   환자 PAT_ID 확인 실패 — 수동 확인 필요. 건너뜀")
                    append_progress(key, None, "no_pat_id")
                    continue
                log(f"   PAT_ID = {pat_id}")

                # 모든 영역을 빈칸이어도 저장하고 넘어감(건너뛰지 않음).
                for area_key, sheet, path in want_areas:
                    vals = area_data.get(area_key, {}).get(key)
                    target = f"{conf['base_url']}{path}?PAT_ID={pat_id}"
                    driver.get(target)
                    try:
                        WebDriverWait(driver, 20).until(
                            EC.presence_of_element_located((By.NAME, "PAT_ID")))
                    except TimeoutException:
                        log(f"   [{area_key}] 페이지 로딩 실패")
                        continue
                    n = fill_form1(driver, resolver, area_key, vals, conf) if vals else 0
                    log(f"   [{area_key}] {n}개 입력" + (" (빈 저장)" if not vals else ""))
                    # 완료체크(핵심변수): comment 와 '빈 y_child(소아 아님)'만 생략, 나머지는 무조건.
                    if not (area_key == "comment" or (area_key == "y_child" and not vals)):
                        mark_complete(driver)
                    save_form1(driver, conf)
                append_progress(key, pat_id, "done")
                log(f"=== 환자키 {key} 완료 (PAT_ID={pat_id}) ===")
            except Exception as e:
                log(f"   오류(환자키 {key}): {e}")
                append_progress(key, None, f"error:{e}")
                dismiss_alert(driver)
        log("전체 처리 종료")
    finally:
        if conf.get("wait_on_finish") and not conf["headless"]:
            try:
                input("끝났습니다. 브라우저를 닫으려면 Enter...")
            except EOFError:
                pass
        try:
            driver.quit()
        except Exception:
            pass


def main():
    if not SELENIUM_OK:
        log("selenium 이 설치되어 있지 않습니다.  설치:  uv sync  (그리고 'uv run python kocarc_bot.py' 로 실행)")
        sys.exit(1)
    conf = load_config()
    conf["wait_on_finish"] = True
    run_bot(conf)


if __name__ == "__main__":
    main()
